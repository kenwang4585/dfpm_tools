
import matplotlib.pyplot as plt
import getpass

from smartsheet_handler import SmartSheetClient
import smartsheet
import re
import openpyxl
import time
#import xlsxwriter # used to avoid
from blg_settings import *
from sending_email import send_attachment_and_embded_image
from db_read import read_table
import gc
import numpy as np
import pandas as pd
import ssl
ssl._create_default_https_context = ssl._create_unverified_context #zzs


plt.rcParams.update({'figure.max_open_warning': 0})


def download_and_send_tracker_as_backup(backup_day,login_user):
    """
    Download and send the addressable backlog files to Ken's email on specified day when glo app runs.
    :return:
    """
    today_name = pd.Timestamp.today().day_name()
    file_list=os.listdir(base_dir_tracker)
    if today_name == backup_day:
        # 添加文件到附件列表 # List of tuples (path, file_name)
        att_files = []
        for file in file_list:
            ext=os.path.splitext(file)[1]
            if ext=='.xlsx':
                att_files.append((base_dir_tracker,file))
        att_files.append((base_dir_db,'foo.db')) # add the db file

        # Send to ken
        to_address = [super_user + '@cisco.com']
        subject = 'DFPM auto tool tracker files - triggered by: {}'.format(login_user)
        html = 'tracker_email.html'

        msg, size_over_limit = send_attachment_and_embded_image(to_address, subject, html,
                                                                att_filenames=att_files,
                                                                bcc=None,
                                                                embeded_filenames=None,)


def add_outlier_col(df=None):
    '''
    Calculate the days for outlier items
    :param df:
    :return:
    '''
    # add entered_not_booked days col
    df.loc[:, 'entered_not_booked'] = np.where(df.BOOKED_DATE.isnull(),
                                               pd.Timestamp.now() - df.LINE_CREATION_DATE, np.nan)
    df.loc[:, 'entered_not_booked'] = df.entered_not_booked.map(lambda x: x.days if x != np.nan else x)
    # add booked_nots_scheduled days col
    df.loc[:, 'booked_not_scheduled'] = np.where(((df.ORIGINAL_FCD_NBD_DATE.isnull()) & (~df.BOOKED_DATE.isnull())),
                                                 pd.Timestamp.now() - df.BOOKED_DATE, np.nan)
    df.loc[:, 'booked_not_scheduled'] = df.booked_not_scheduled.map(lambda x: x.days if x != np.nan else x)
    # add missing OSSD days col
    df.loc[:, 'missed_ossd'] = np.where(((df.ORIGINAL_FCD_NBD_DATE.notnull()) & (df.PACKOUT_QUANTITY != 'Packout Completed')),
                                    np.where(pd.Timestamp.now() > df.ORIGINAL_FCD_NBD_DATE,
                                             (pd.Timestamp.now() - df.ORIGINAL_FCD_NBD_DATE),
                                             np.nan),
                                    np.nan)
    df.loc[:, 'missed_ossd'] = df.missed_ossd.map(lambda x: x.days if x != np.nan else x)

    # missed recommit
    df.loc[:,'missed_recommit']=np.where((df.CURRENT_FCD_NBD_DATE.notnull())&(df.CURRENT_FCD_NBD_DATE<pd.Timestamp.today()) &
                                         (df.ADDRESSABLE_FLAG!='PO_CANCELLED') & (df.PACKOUT_QUANTITY!='Packout Completed'),
                                         pd.Timestamp.today()-df.CURRENT_FCD_NBD_DATE,
                                         np.nan)
    df.loc[:, 'missed_recommit'] =df.missed_recommit.map(lambda x: x.days if x != np.nan else x)

    df=add_cancel_cols(df)
    df=add_partial_staged_col(df)

    return df


def add_partial_staged_col(df):
    '''
    Calculate the aging days for partial staged SS based on the longest aging days PO since ASN_CREATION_DATE
    also ad the unstaged PO columns and also aggregated comments from each PO comments
    :param df: df (ASN_CREATION_DATE must be in Date format)
    :return: df
    '''
    # below to find out PO under partial staged SS:
    x = zip(df.SO_SS, df.PO_NUMBER, df.ASN_CREATION_DATE)
    d = [[x, y, str(z)] for x, y, z in x]
    c = np.array(d)
    f = np.unique(c, axis=0)

    g = {}  # create a dict{SO_SS:[[po,asn]]}
    for row in f:
        po_asn = [[row[1], row[2]]]
        if row[0] in g.keys():
            po_asn = g.get(row[0]) + po_asn

        g[row[0]] = po_asn

    for key in list(g):  # remove the single line SS... remaining SS would be multiple PO SS
        if len(g[key]) == 1:
            g.pop(key)

    for key in list(g):  # remove SS that have all PO asn'ed
        all_asn = False
        for group in g[key]:
            if 'NaT' in group:
                all_asn = True
                continue
        if all_asn == False:
            g.pop(key)

    for key in list(g):  # remove SS that have None PO asn'ed
        none_asn = True
        for group in g[key]:
            if 'NaT' not in group:
                none_asn = False
                continue
        if none_asn == True:
            g.pop(key)

    # generate partial asn'ed PO list (keep the asn'ed and remove the non-asn'ed PO)
    partial_po = [y[0] for x in g.values() for y in x if y[1] != 'nan']

    # print(partial_po)

    # below based on partial_po to to calculate min_date of ASN under each partial staged SS:
    zip_data = zip(df.SO_SS, df.PO_NUMBER, df.ASN_CREATION_DATE, df.C_UNSTAGED_DOLLARS, df.BUSINESS_UNIT,
                   df.COMMENTS)
    d = [[x, y, z, w, p, q] for x, y, z, w, p, q in zip_data]

    # print(d)

    # 根据已经在partial_po下的PO计算最小asn date及ss revenue,并创建字典：{SS:[min_date,ss_rev]}
    g = {}
    for group in d:
        if group[1] in partial_po:
            if group[0] in g.keys():
                min_date = min(g[group[0]][0], group[2])
                # ss_rev=g[group[0]][1]+group[3]
            else:
                min_date = group[2]
                # ss_rev=group[3]

            g[group[0]] = [min_date]

    # 根据partial_po下的PO统计SS下还没有AS_CREATION_DATE的PO list,以及从po_number下抓取comments
    x = {}
    for group in d:
        if group[1] in partial_po:
            if pd.isnull(group[2]):  # ASN_CREATION_DATE is NaT
                if group[0] in x.keys():  # 如果已经加入字典
                    if group[1] + '(' + str(group[5]) + ')' not in x[group[0]][0]:  # PO(BU)/PO(BU)
                        x[group[0]][0] += '/ ' + group[1] + '(' + group[4] + ')'  # PO(BU)/PO(BU)
                        x[group[0]][1] += '/ ' + group[1] + '(' + group[4] + ': ' + str(group[5]) + ')\n'
                else:
                    x[group[0]] = ['', '']
                    x[group[0]][0] = group[1] + '(' + group[4] + ')'
                    x[group[0]][1] = group[1] + '(' + group[4] + ': ' + str(group[5]) + ')\n'

    # print(x)

    # calculate staging days for partial SS based on min_date of ASN
    df.loc[:, 'ss_partial_staged_days'] = df.SO_SS.map(
        lambda x: pd.Timestamp.now() - g.get(x)[0] if x in g.keys() else np.nan)
    df.loc[:, 'ss_partial_staged_days'] = df.ss_partial_staged_days.map(lambda x: x.days if x != np.nan else x)
    # df.loc[:,'ss_rev']=df.SO_SS.map(lambda x: g.get(x)[1] if x in g.keys() else np.nan)

    # add col to indicate which PO is not yet staged - ASN_CREATION_DATE
    df.loc[:, 'unstaged_po'] = df.SO_SS.map(lambda y: x[y][0] if y in x.keys() else np.nan) #for summary
    df.loc[:, 'unstaged_under_ss']=np.where((df.ASN_CREATION_DATE.isnull())&(df.unstaged_po.notnull()), # for 3a4
                                      'YES',
                                      None)
    # change comments to combined comments
    df.loc[:, 'partial_comments'] = df.SO_SS.map(lambda y: x[y][1] if y in x.keys() else np.nan)

    del zip_data
    gc.collect()

    return df

def add_cancel_cols(df):
    '''
    Add 'need_rtv' and 'cancel_aging_days' indication for cancelled order.
    col 'need_rtv' is based on if PACKOUT_QUANTITY is null; 'cancel_aging_days' is not based on real cancellation hold
    days, but based on below sequence: ASN_CREATION_DATE->LT_TARGET_FCD.
    :param df: df
    :return: df
    '''

    # add Need_RTV col
    #df.loc[:, 'need_rtv'] = np.nan
    df.loc[:, 'need_rtv'] = np.where(df.ADDRESSABLE_FLAG == 'PO_CANCELLED',
                                     np.where(df.PACKOUT_QUANTITY.notnull(),
                                              'YES',
                                              'NO'),
                                     None)
    # add aging days col
    df.loc[:, 'cancel_aging_days'] = np.where(df.ADDRESSABLE_FLAG == 'PO_CANCELLED',
                                              np.where(df.ASN_CREATION_DATE.notnull(),
                                                       pd.Timestamp.now() - df.ASN_CREATION_DATE,
                                                       np.where(df.LT_TARGET_FCD.notnull(),
                                                                np.where(pd.Timestamp.now() > df.LT_TARGET_FCD,
                                                                        pd.Timestamp.now() > df.LT_TARGET_FCD,
                                                                        np.nan),
                                                                0)),
                                              np.nan)

    df.loc[:, 'cancel_aging_days'] = df.cancel_aging_days.map(lambda x: x.days if x != np.nan else x)

    return df


def read_subscription_by_region():
    """
    Read the subscrition db for emails by region
    """
    df_subscription = read_table('subscription')

    backlog_dashboard_emails_global={}
    wnbu_compliance_check_emails_global={}
    config_check_emails_global={}
    blg_apjc,blg_emea,blg_americas=[],[],[]
    wnbu_apjc, wnbu_emea, wnbu_americas = [], [], []
    config_apjc, config_emea, config_americas = [], [], []

    for row in df_subscription.itertuples():
        email=row.Email
        tasks=eval(row.Subscription)
        for key,values in tasks.items():
            if key=='Backlog dashboard':
                for value in values:
                    if value=='APJC':
                        blg_apjc.append(email)
                    elif value=='EMEA':
                        blg_emea.append(email)
                    else:
                        blg_americas.append(email)
            elif key=='WNBU compliance':
                for value in values:
                    if value=='APJC':
                        wnbu_apjc.append(email)
                    elif value=='EMEA':
                        wnbu_emea.append(email)
                    else:
                        wnbu_americas.append(email)
            elif key=='Config report':
                for value in values:
                    if value=='APJC':
                        config_apjc.append(email)
                    elif value=='EMEA':
                        config_emea.append(email)
                    else:
                        config_americas.append(email)

    backlog_dashboard_emails_global['APJC']=blg_apjc
    backlog_dashboard_emails_global['EMEA']=blg_emea
    backlog_dashboard_emails_global['Americas']=blg_americas

    wnbu_compliance_check_emails_global['APJC']=wnbu_apjc
    wnbu_compliance_check_emails_global['EMEA']=wnbu_emea
    wnbu_compliance_check_emails_global['Americas']=wnbu_americas

    config_check_emails_global['APJC']=config_apjc
    config_check_emails_global['EMEA']=config_emea
    config_check_emails_global['Americas']=config_americas

    return backlog_dashboard_emails_global,wnbu_compliance_check_emails_global,config_check_emails_global

def read_subscription_by_site(org):
    """
    Read the subscrition db for emails by org code
    """
    df_subscription = read_table('subscription')

    ranking_cm =[] # based on specific org

    for row in df_subscription.itertuples():
        email=row.Email
        tasks=eval(row.Subscription)
        for key,values in tasks.items():
            if key=='Backlog ranking':
                for value in values:
                    if value==org:
                        ranking_cm.append(email)

    cm_emails=ranking_cm

    return cm_emails


def commonize_and_create_main_item(df, col, new_col):
    '''
    Commonize values based on the main item value
    :param df:
    :param col: Col to refer to
    :param new_col: New created col
    :return: df
    '''
    if 'OPTION_NUMBER' in df.columns:
        df_main = df[df.OPTION_NUMBER == 0]
    else:
        df_main = df

    main_dic = {}

    for value, po in zip(df_main[col], df_main.PO_NUMBER.values):
        main_dic[po] = value

    df.loc[:, new_col] = df.PO_NUMBER.map(lambda x: main_dic[x] if x in main_dic else 'Missing Option 0')

    return df

def decide_qend_date(qend_list):
    """
    Decide which date is curent qend date based on predefined qend_list.
    :param qend_list:
    :return:
    """
    qend_list=pd.to_datetime(qend_list)
    today=pd.Timestamp.today().date()

    for dt in qend_list:
        if dt>today:
            qend=dt
            break

    return qend

def create_po_rev_unstg_col(df_3a4):
    """
    Create po rev col for 3a4 with option items
    :param df_3a4:
    :return:
    """
    ### Step0: 计算po_rev - not used in ranking but other calculations
    po_rev_unstg = {}
    df_rev = df_3a4.pivot_table(index='PO_NUMBER', values='C_UNSTAGED_DOLLARS', aggfunc=sum)
    for po, rev in zip(df_rev.index, df_rev.values):
        po_rev_unstg[po] = rev[0]
    df_3a4.loc[:, 'po_rev_unstg'] = df_3a4.PO_NUMBER.map(lambda x: po_rev_unstg[x])

    return df_3a4


def read_backlog_priority_from_smartsheet(df_3a4,login_user):
    '''
    Read backlog priorities from smartsheet; remove SS showing packed/cancelled, or created by self but disappear from 34(if the org/BU also exist in 3a4.);
     create and segregate to top priority and mid priority
    :return:
    '''
    # 从smartsheet读取backlog
    token = os.getenv('PRIORITY_TOKEN')
    sheet_id = os.getenv('PRIORITY_ID')
    proxies = None  # for proxy server
    smartsheet_client = SmartSheetClient(token, proxies)
    df_smart = smartsheet_client.get_sheet_as_df(sheet_id, add_row_id=True, add_att_id=False)

    # Identify SS not in df_3a4 that can be removed - SS created by self and is disappeared from 3a4 - if the 3a4 include the org and BU
    df_smart_self=df_smart[df_smart['Created By']==login_user+'@cisco.com']
    df_smart_w_org_bu_in_3a4 = df_smart_self[
        (df_smart_self.ORG.isin(df_3a4.ORGANIZATION_CODE.unique())) & (df_smart_self.BU.isin(df_3a4.BUSINESS_UNIT.unique()))]
    ss_not_in_3a4 = np.setdiff1d(df_smart_w_org_bu_in_3a4.SO_SS.values, df_3a4.SO_SS.values)

    # SS showing as packed or cancelled in 3a4
    ss_cancelled_or_packed_3a4 = get_packed_or_cancelled_ss_from_3a4(df_3a4)

    # total ss to remove
    df_removal = df_smart[(df_smart.SO_SS.isin(ss_cancelled_or_packed_3a4)) | (df_smart.SO_SS.isin(ss_not_in_3a4))]

    # create the priority dict
    df_smart.drop_duplicates('SO_SS', keep='last', inplace=True)
    df_smart = df_smart[(df_smart.SO_SS.notnull()) & (df_smart.Ranking.notnull())]
    ss_exceptional_priority = {}
    priority_top = {}
    priority_mid = {}
    for row in df_smart.itertuples():
        try: # in case error input of non-num ranking
            if float(row.Ranking)<4:
                priority_top[row.SO_SS] = float(row.Ranking)
            else:
                priority_mid[row.SO_SS] = float(row.Ranking)
        except:
            print('{} has a wrong ranking#: {}.'.format(row.SO_SS,row.Ranking) )

        ss_exceptional_priority['priority_top'] = priority_top
        ss_exceptional_priority['priority_mid'] = priority_mid

    return ss_exceptional_priority,df_removal

def remove_priority_ss_from_smtsheet_and_notify(df_removal,login_user,sender='APJC DFPM'):
    """
    Remove the packed/cancelled SS from priority smartsheet and send email to corresponding people for whose SS are removed from the priority smartsheet
    """
    if df_removal.shape[0]>0:
        token = os.getenv('PRIORITY_TOKEN')
        sheet_id = os.getenv('PRIORITY_ID')
        proxies = None  # for proxy server
        smartsheet_client = SmartSheetClient(token, proxies)

        removal_row_id = df_removal.row_id.values.tolist()
        removal_ss_email = list(set(df_removal['Created By'].values.tolist()))
        if len(removal_row_id) > 0:
            smartsheet_client.delete_row(sheet_id=sheet_id, row_id=removal_row_id)

        to_address = removal_ss_email
        to_address = to_address + [login_user+'@cisco.com']
        html_template='priority_ss_removal_email.html'
        subject='SS auto removal from exceptional priority smartsheet - by {}'.format(login_user)

        send_attachment_and_embded_image(to_address, subject, html_template, att_filenames=None,
                                         embeded_filenames=None,
                                         sender=sender,
                                         bcc=None,
                                         removal_ss_header=df_removal.columns,
                                         removal_ss_details=df_removal.values,
                                         user=login_user)




def get_packed_or_cancelled_ss_from_3a4(df_3a4):
    """
    Get the fully packed or canceleld SS from 3a4 - for deleting exceptional priority smartsheet purpose.
    """
    ss_cancelled=df_3a4[df_3a4.ADDRESSABLE_FLAG=='PO_CANCELLED'].SO_SS.unique()

    ss_with_po_packed=df_3a4[df_3a4.PACKOUT_QUANTITY=='Packout Completed'].SO_SS.unique()
    ss_wo_po_packed = df_3a4[df_3a4.PACKOUT_QUANTITY != 'Packout Completed'].SO_SS.unique() # some PO may not be packed in one SS
    ss_fully_packed=np.setdiff1d(ss_with_po_packed,ss_wo_po_packed)

    ss_packed_not_cancelled=np.setdiff1d(ss_fully_packed,ss_cancelled)

    ss_cancelled_or_packed_3a4=ss_cancelled.tolist()+ss_packed_not_cancelled.tolist()

    return ss_cancelled_or_packed_3a4


def get_file_info_on_drive(base_path,keep_hours=100):
    """
    Collect the file info on a drive and make that into a df. Remove files if older than keep_hours.
    """
    now=time.time()
    file_list = os.listdir(base_path)
    if '.keep' in file_list:
        file_list.remove('.keep')

    files = []
    creation_time = []
    file_size = []
    file_path = []
    for file in file_list:
        c_time = os.stat(os.path.join(base_path, file)).st_ctime

        if (now - c_time) / 3600 > keep_hours: #hours
            os.remove(os.path.join(base_path, file))
        else:
            c_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(c_time))
            file_s = os.path.getsize(os.path.join(base_path, file))
            if file_s > 1024 * 1024:
                file_s = str(round(file_s / (1024 * 1024), 1)) + 'M'
            else:
                file_s = str(int(file_s / 1024)) + 'K'

            files.append(file)
            creation_time.append(c_time)
            file_size.append(file_s)
            file_path.append(os.path.join(base_path,file))

    df_file_info=pd.DataFrame({'File_name':files,'Creation_time':creation_time, 'File_size':file_size, 'File_path':file_path})
    df_file_info.sort_values(by='Creation_time',ascending=False,inplace=True)

    return df_file_info


#  Depracated ----- changed to below Jan version which include riso_ranking
def ss_ranking_overall_new_december(df_3a4,ss_exceptional_priority,ranking_col, lowest_priority_cat,order_col='SO_SS', new_col='ss_overall_rank'):
    """
    按照ranking_col的顺序对SS进行排序。最后放MFG_HOLD订单.
    注：CTB和PCBA allocation用相同的方式在开始处删除cancelled的订单；summary_3a4不删除cancelled订单，不过在结尾处清除cancelled订单的ranking#
    """
     # Below create a rev_rank for reference -  currently not used in overall ranking
    ### change non-rev orders unstaged $ to 0
    df_3a4.loc[:,'C_UNSTAGED_DOLLARS']=np.where(df_3a4.REVENUE_NON_REVENUE == 'NO',
                                                0,
                                                df_3a4.C_UNSTAGED_DOLLARS)

    #### 生成ss_unstg_rev - 在这里不参与排序
    # 计算ss_unstg_rev
    ss_unstg_rev = {}
    df_rev = df_3a4.pivot_table(index='SO_SS', values='C_UNSTAGED_DOLLARS', aggfunc=sum)
    for ss, rev in zip(df_rev.index, df_rev.values):
        ss_unstg_rev[ss] = rev[0]
    df_3a4.loc[:, 'ss_unstg_rev'] = df_3a4.SO_SS.map(lambda x: ss_unstg_rev[x])

    """
    # 计算po_rev_unit - non revenue change to 0
    df_3a4.loc[:, 'po_rev_unit'] = np.where(df_3a4.REVENUE_NON_REVENUE == 'YES',
                                            df_3a4.SOL_REVENUE / df_3a4.ORDERED_QUANTITY,
                                            0)

    # 计算ss_rev_unit: 通过po_rev_unit汇总
    ss_rev_unit = {}
    dfx_rev = df_3a4.pivot_table(index='SO_SS', values='po_rev_unit', aggfunc=sum)
    for ss, rev in zip(dfx_rev.index, dfx_rev.values):
        ss_rev_unit[ss] = rev[0]
    df_3a4.loc[:, 'ss_rev_unit'] = df_3a4.SO_SS.map(lambda x: int(ss_rev_unit[x]))
    """

    # create rank#
    rank = {}
    order_list = df_3a4.sort_values(by='ss_unstg_rev', ascending=False).SO_SS.unique()
    for order, rk in zip(order_list, range(1, len(order_list) + 1)):
        rank[order] = rk
    df_3a4.loc[:, 'ss_rev_rank'] = df_3a4.SO_SS.map(lambda x: rank[x])

    # below creates overall ranking col
    ### Step1: 重新定义priority order及排序
    df_3a4.loc[:, 'priority_cat'] = np.where(df_3a4.SECONDARY_PRIORITY.isin(['PR1', 'PR2', 'PR3']),
                                             df_3a4.SECONDARY_PRIORITY,
                                             np.where(df_3a4.FINAL_ACTION_SUMMARY == 'LEVEL 4 ESCALATION PRESENT',
                                                      'L4',
                                                      np.where(df_3a4.BUP_RANK.notnull(),
                                                               'BUP',
                                                                np.where(df_3a4.PROGRAM.notnull(),
                                                                        'YE',
                                                                         None))))

    #### Update below DX/DO orders to PR1/PR2 due to current PR1/2/3 not updated when order change to DPAS from others
    df_3a4.loc[:, 'priority_cat']=np.where((df_3a4.DPAS_RATING.isin(['DX','TAA-DX']))&(df_3a4.priority_cat.isnull()),
                                           'PR1',
                                           df_3a4.priority_cat)
    df_3a4.loc[:, 'priority_cat'] = np.where((df_3a4.DPAS_RATING.isin(['DO', 'TAA-DO'])) & (df_3a4.priority_cat.isnull()),
                                            'PR2',
                                            df_3a4.priority_cat)

    #### Step2: Generate rank for priority orders
    df_3a4.loc[:, 'priority_rank_top'] = np.where(df_3a4.priority_cat == 'PR1',
                                              1,
                                              np.where(df_3a4.priority_cat == 'PR2',
                                                       2,
                                                       np.where(df_3a4.priority_cat == 'PR3',
                                                                3,
                                                                None)))

    df_3a4.loc[:, 'priority_rank_mid'] =np.where(df_3a4.priority_cat == 'L4',
                                            4,
                                            np.where(df_3a4.priority_cat == 'BUP',
                                                    5,
                                                    np.where(df_3a4.priority_cat == 'YE',
                                                             6,
                                                             None)))

    #### update ranking based on exception priority setting
    df_3a4.loc[:, 'priority_rank_top'] = np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_top'].keys()),
                                                  df_3a4.SO_SS.map(lambda x: ss_exceptional_priority['priority_top'].get(x)),
                                                  np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_mid'].keys()),
                                                            None,
                                                            df_3a4.priority_rank_top))
    df_3a4.loc[:, 'priority_rank_mid'] = np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_mid'].keys()),
                                                  df_3a4.SO_SS.map(lambda x: ss_exceptional_priority['priority_mid'].get(x)),
                                                  np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_top'].keys()),
                                                            None,
                                                            df_3a4.priority_rank_mid))


    # Create a new col to indicate the rank - in ranking, actually use priority_rank_top and priority_rank_mid
    df_3a4.loc[:, 'priority_rank'] = np.where(df_3a4.priority_rank_top.notnull(),
                                              df_3a4.priority_rank_top,
                                              df_3a4.priority_rank_mid)


    ##### Step3: Give revenue/non-revenue a rank
    df_3a4.loc[:,'rev_non_rev_rank']=np.where(df_3a4.REVENUE_NON_REVENUE=='YES', 0, 1)

    ##### Step4: sort the SS per ranking columns and Put MFG hold orders at the back
    df_3a4.sort_values(by=ranking_col, ascending=True, inplace=True)

    # Put lowest priority orders at the back -- referring lowest_priority_cat in setting
    # !!!! different from CTB&pcba allocation
    df_low_prority=pd.DataFrame()
    for col_name, col_val in lowest_priority_cat.items():
        dfx=df_3a4[df_3a4[col_name].isin(col_val)]
        df_low_prority=pd.concat([df_low_prority,dfx],sort=False)
        df_3a4=df_3a4[~df_3a4[col_name].isin(col_val)].copy()
    df_3a4 = pd.concat([df_3a4, df_low_prority], sort=False)

    ##### Step5: create rank# and put in 3a4
    rank = {}
    order_list = df_3a4[order_col].unique()
    for order, rk in zip(order_list, range(1, len(order_list) + 1)):
        rank[order] = rk
    df_3a4.loc[:, new_col] = df_3a4[order_col].map(lambda x: rank[x])

    ##### Step6: !!! unique for summary_3a4 - remove rank for cancelled orders
    df_3a4.loc[:, new_col]=np.where(df_3a4.ADDRESSABLE_FLAG=='PO_CANCELLED',
                                    None,
                                    df_3a4[new_col])

    return df_3a4



def create_dfpm_3a4(dfpm_mapping,dfpm_3a4_option,df_3a4,df_asp,addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf, ctb_summary_for_material,login_user):
    '''
    Create 3a4 file for DFPM
    '''

    """
    # By Org: 生成backlog distribution summary based on date labels (CRD, target LT FCD, current FCD,OSSD)
    df_backlog_distribution_qty_org = create_backlog_distribution_summary_by_org_new(df_3a4,
                                                                                     backlog_distribution_date_labels,
                                                                                     data_col='C_UNSTAGED_QTY',
                                                                                     func='SUM', by='PO_NUMBER',
                                                                                     excl_packed=True)
    df_backlog_distribution_amt_org = create_backlog_distribution_summary_by_org_new(df_3a4,
                                                                                     backlog_distribution_date_labels,
                                                                                     data_col='po_unstaged_rev',
                                                                                     func='SUM', by='PO_NUMBER',
                                                                                     excl_packed=True)
    df_backlog_distribution_po_count_org = create_backlog_distribution_summary_by_org_new(df_3a4,
                                                                                          backlog_distribution_date_labels,
                                                                                          data_col='PO_NUMBER',
                                                                                          func='COUNT',
                                                                                          by='PO_NUMBER',
                                                                                          excl_packed=True)
    df_backlog_distribution_ss_count_org = create_backlog_distribution_summary_by_org_new(df_3a4,
                                                                                          backlog_distribution_date_labels,
                                                                                          data_col='SO_SS',
                                                                                          func='COUNT',
                                                                                          by='SO_SS',
                                                                                          excl_packed=True)
    """


    df_dict={}
    df_3a4=df_3a4[col_3a4_dfpm]
    df_3a4.set_index('ORGANIZATION_CODE', inplace=True)
    dfpm_3a4_created=[]

    #
    file_path = base_dir_output
    df_dict['addr_ctb_bu'] = addr_ctb_by_org_bu
    df_dict['addr_ctb_pf'] = addr_ctb_by_org_bu_pf
    df_dict['ctb_material'] = ctb_summary_for_material
    df_dict['asp'] = df_asp

    if dfpm_3a4_option=='by_dfpm':
        #create total 3a4 with option==0 only
        dfy=df_3a4[df_3a4.OPTION_NUMBER == 0].copy()
        dfy.drop(['C_UNSTAGED_DOLLARS'],axis=1,inplace=True)
        df_dict['3a4'] = dfy
        fname = 'Site 3a4 for main PID (created by ' + login_user + ') ' + pd.Timestamp.now().strftime(
            '%m-%d %H:%M') + '.xlsx'
        full_path = os.path.join(file_path, fname)
        write_excel_file(full_path, df_dict)
        dfpm_3a4_created.append('site level')

        # create file by DFPM
        for dfpm,coverage in dfpm_mapping.items():
            df_dfpm=pd.DataFrame()
            for org,contents in coverage.items():
                if org in df_3a4.index:
                    dfy=df_3a4.loc[org].copy()
                    dfy=dfy[((dfy.main_bu.isin(contents[0])&(~dfy.main_pf.isin(contents[2])))|(dfy.main_pf.isin(contents[1])))].copy()
                    if dfy.shape[0]>0:
                        df_dfpm=pd.concat([df_dfpm,dfy],sort=False)

                    del dfy

            if df_dfpm.shape[0]>0:
                df_dict['3a4'] = df_dfpm

                fname = dfpm + ' 3a4 (created by ' + login_user + ') ' + pd.Timestamp.now().strftime('%m-%d %H:%M') +'.xlsx'
                full_path = os.path.join(file_path, fname)
                write_excel_file(full_path, df_dict)
                dfpm_3a4_created.append(dfpm)

                to_address = [dfpm + '@cisco.com']
                subject = '3a4 summary for {}'.format(dfpm)
                html = 'dfpm_3a4.html'

                msg, size_over_limit = send_attachment_and_embded_image(to_address, subject, html,
                                                                    att_filenames=None,
                                                                    file_3a4=fname)

        del df_dfpm
        gc.collect()
    else:
        df_dict['3a4'] = df_3a4
        fname = '3a4 (created by ' + login_user + ') ' + pd.Timestamp.now().strftime('%m-%d %H:%M') + '.xlsx'
        full_path = os.path.join(file_path, fname)
        write_excel_file(full_path, df_dict)
        dfpm_3a4_created.append(login_user)

    return dfpm_3a4_created

# newly introduced in Jan 2021 to add riso_ranking
def ss_ranking_overall_new_jan(df_3a4,ss_exceptional_priority,ranking_col_dic, lowest_priority_cat,order_col='SO_SS',with_dollar=True):
    """
    按照ranking_col的顺序对SS进行排序。最后放MFG_HOLD订单.
    注：CTB和PCBA allocation用相同的方式在开始处删除cancelled的订单；summary_3a4不删除cancelled订单，不过在结尾处清除cancelled订单的ranking#
    """
    if with_dollar:
        # Below create a rev_rank for reference -  currently not used in overall ranking
        ### change non-rev orders unstaged $ to 0
        df_3a4.loc[:,'C_UNSTAGED_DOLLARS']=np.where(df_3a4.REVENUE_NON_REVENUE == 'NO',
                                                    0,
                                                    df_3a4.C_UNSTAGED_DOLLARS)

        #### 生成ss_unstg_rev - 在这里不参与排序
        # 计算ss_unstg_rev
        ss_unstg_rev = {}
        df_rev = df_3a4.pivot_table(index='SO_SS', values='C_UNSTAGED_DOLLARS', aggfunc=sum)
        for ss, rev in zip(df_rev.index, df_rev.values):
            ss_unstg_rev[ss] = rev[0]
        df_3a4.loc[:, 'ss_unstg_rev'] = df_3a4.SO_SS.map(lambda x: ss_unstg_rev[x])

        # create rank#
        rank = {}
        order_list = df_3a4.sort_values(by='ss_unstg_rev', ascending=False).SO_SS.unique()
        for order, rk in zip(order_list, range(1, len(order_list) + 1)):
            rank[order] = rk
        df_3a4.loc[:, 'ss_rev_rank'] = df_3a4.SO_SS.map(lambda x: rank[x])

    ### Step 1-1: 重新定义priority order及排序
    df_3a4.loc[:, 'priority_cat'] = np.where(df_3a4.SECONDARY_PRIORITY.isin(['PR1', 'PR2', 'PR3']),
                                                 df_3a4.SECONDARY_PRIORITY,
                                                 np.where(df_3a4.FINAL_ACTION_SUMMARY == 'LEVEL 4 ESCALATION PRESENT',
                                                          'L4',
                                                          np.where(df_3a4.BUP_RANK.notnull(),
                                                                   'BUP',
                                                                   None)))

    #### Update below DX/DO orders to PR1/PR2 due to current PR1/2/3 not updated when order change to DPAS from others
    df_3a4.loc[:, 'priority_cat']=np.where((df_3a4.DPAS_RATING.isin(['DX','TAA-DX']))&(df_3a4.priority_cat.isnull()),
                                           'PR1',
                                           df_3a4.priority_cat)
    df_3a4.loc[:, 'priority_cat'] = np.where((df_3a4.DPAS_RATING.isin(['DO', 'TAA-DO'])) & (df_3a4.priority_cat.isnull()),
                                            'PR2',
                                            df_3a4.priority_cat)

    #### Step 1-2: Generate rank for priority orders
    df_3a4.loc[:, 'priority_rank_top'] = np.where(df_3a4.priority_cat == 'PR1',
                                              1,
                                              np.where(df_3a4.priority_cat == 'PR2',
                                                       2,
                                                       np.where(df_3a4.priority_cat == 'PR3',
                                                                3,
                                                                None)))

    df_3a4.loc[:, 'priority_rank_mid'] =np.where(df_3a4.priority_cat == 'L4',
                                            4,
                                            np.where(df_3a4.priority_cat == 'BUP',
                                                    5,
                                                     None))

    #### update ranking based on exception priority setting
    df_3a4.loc[:, 'priority_rank_top'] = np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_top'].keys()),
                                                  df_3a4.SO_SS.map(lambda x: ss_exceptional_priority['priority_top'].get(x)),
                                                  np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_mid'].keys()),
                                                            None,
                                                            df_3a4.priority_rank_top))
    df_3a4.loc[:, 'priority_rank_mid'] = np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_mid'].keys()),
                                                  df_3a4.SO_SS.map(lambda x: ss_exceptional_priority['priority_mid'].get(x)),
                                                  np.where(df_3a4.SO_SS.isin(ss_exceptional_priority['priority_top'].keys()),
                                                            None,
                                                            df_3a4.priority_rank_mid))


    # Create a new col to indicate the rank based on setting - in ranking, actually use priority_rank_top and priority_rank_mid
    df_3a4.loc[:, 'priority_rank'] = np.where(df_3a4.priority_rank_top.notnull(),
                                              df_3a4.priority_rank_top,
                                              df_3a4.priority_rank_mid)


    ##### Step 1-3: Give revenue/non-revenue a rank
    df_3a4.loc[:,'rev_non_rev_rank']=np.where(df_3a4.REVENUE_NON_REVENUE=='YES', 0, 1)



    ##### 2: Create the overall_rank (or called Prod_ranking)
    for ranking_name,ranking_col in ranking_col_dic.items():
        ##### Step 2-1: sort the SS per ranking columns
        df_3a4.sort_values(by=ranking_col, ascending=True, inplace=True)

        # Put lowest priority orders at the back -- referring lowest_priority_cat in setting
        # !!!! different from CTB&pcba allocation
        df_low_prority=pd.DataFrame()
        for col_name, col_val in lowest_priority_cat.items():
            dfx=df_3a4[df_3a4[col_name].isin(col_val)]
            df_low_prority=pd.concat([df_low_prority,dfx],sort=False)
            df_3a4=df_3a4[~df_3a4[col_name].isin(col_val)].copy()
        df_3a4 = pd.concat([df_3a4, df_low_prority], sort=False)

        ##### Step 2-2: create rank# and put in 3a4
        rank = {}
        order_list = df_3a4[order_col].unique()
        for order, rk in zip(order_list, range(1, len(order_list) + 1)):
            rank[order] = rk
        df_3a4.loc[:, ranking_name] = df_3a4[order_col].map(lambda x: rank[x])

        ##### Step 2-3: !!! unique for summary_3a4 - remove rank for cancelled orders
        df_3a4.loc[:, ranking_name]=np.where(df_3a4.ADDRESSABLE_FLAG=='PO_CANCELLED',
                                        None,
                                        df_3a4[ranking_name])


    return df_3a4

def create_addr_summary(df, org_name_region):
    '''
    Create addressable summary based on df_3a4 and org_name map
    '''

    addr_df_summary = []
    for name,orgs in org_name_region.items():
        dfx = df[df.ORGANIZATION_CODE.isin(orgs)].pivot_table(index='main_bu',
                                                                  columns=['ADDRESSABLE_FLAG'],
                                                                  values='po_rev_unstg', aggfunc=sum)

        dfx.loc[:, 'Total'] = dfx.sum(axis=1)
        dfx.loc[name] = dfx.sum(axis=0)

        dfx.columns.name = name
        dfx.index.name = 'BU'
        #dfx.sort_values(by='Total', ascending=False, inplace=True)
        dfx = dfx.applymap(lambda x: round(x / 1000000, 1))

        if 'ADDRESSABLE' in dfx.columns:
            addr_df_summary.append(dfx)

    del dfx
    gc.collect()

    #print(addr_df_summary)

    return addr_df_summary

# Currently not used
def create_blg_summary_per_date_label(df, date_labels, bucket='M'):
    '''
    Create rev summaries based on different date labels
    :param bucket: Can be "M", "W", or "Q" - for the date buckets
    :param df:
    :param date_labels: set in basic_settings
    :return: summary df
    '''

    rev_df_summary = []
    for label in date_labels:
        # just pick out option0 and count the PO numbers
        df = df[df.OPTION_NUMBER == 0].copy()
        dfx = df.pivot_table(index=label, columns='ORGANIZATION_CODE', values='PO_NUMBER', aggfunc=len)
        # print(dfx)

        dfx.loc[:, 'APJC'] = dfx.sum(axis=1)
        col = ['APJC'] + dfx.columns[:-1].tolist()
        dfx = dfx[col]
        dfx = dfx[dfx.APJC > 0].copy()

        # print(dfx)

        # raise Exception('end')

        dfx = dfx.resample(bucket,label='left').sum()
        if bucket == 'M':
            dfx.index = dfx.index.strftime('%y-%m')
        dfx.columns.name = label
        # dfx=dfx.applymap(lambda x: round(x/1000000,1))

        rev_df_summary.append(dfx)

    orgs = rev_df_summary[0].columns

    # 对dfx进行变形处理
    rev_df_summary_new = []
    for org in orgs:
        summary_by_org = []
        for dfx in rev_df_summary:
            dfy = pd.DataFrame({dfx.columns.name: dfx[org]})
            summary_by_org.append(dfy)

        df_c = pd.concat(summary_by_org, axis=1, sort=False)
        df_c.columns.name = org
        '''
        #增加cumsum列
        for col in df_c.columns:
           new_col=col + '_Cumsum'
           df_c.loc[:,new_col]=df_c[col].cumsum()
        '''
        df_c.rename(columns={'CUSTOMER_REQUEST_DATE': 'CRD',
                             'LT_TARGET_FCD': 'LT_FCD',
                             'CURRENT_FCD_NBD_DATE': 'C_FCD'}, inplace=True)

        rev_df_summary_new.append(df_c)

    del dfx, rev_df_summary, df_c, summary_by_org
    gc.collect()

    return rev_df_summary_new




def collect_new_addr_and_comb_historical_apjc(df_list, addr_fname):
    '''
    This program convert the df_list (addressable backlog summary list) formats, mark it with current date for hte latest record. And concat it with what's already in the addressable_tracker.
    :param df_list: Summary created which is a list of df: addr_df_summary
    :param addr_fname: file name of the tracker file
    :return:
    '''
    old_addr_apjc = pd.read_excel(addr_fname, sheet_name='APJC', index_col='DATE')
    old_addr_foc = pd.read_excel(addr_fname, sheet_name='FOC', index_col='DATE')
    old_addr_fdo = pd.read_excel(addr_fname, sheet_name='FDO', index_col='DATE')
    old_addr_shk = pd.read_excel(addr_fname, sheet_name='SHK', index_col='DATE')
    old_addr_jpe = pd.read_excel(addr_fname, sheet_name='JPE', index_col='DATE')
    old_addr_ncb = pd.read_excel(addr_fname, sheet_name='NCB', index_col='DATE')

    # 提取apjc by org数据
    df_apjc = pd.concat(df_list, sort=False)
    df_apjc = df_apjc.loc[['APJC', 'FOC', 'FDO', 'SHK', 'JPE', 'NCB']]
    #print(df_apjc)
    df_apjc = df_apjc.ADDRESSABLE

    # 以下将Series转置并于原使数据合并
    dic = {}
    for org in df_apjc.index:
        dic[org] = df_apjc.loc[org]
    df_apjc = pd.DataFrame(dic, index=[pd.Timestamp.now().strftime('%Y-%m-%d')])
    df_apjc.index.name = 'DATE'
    df_apjc = pd.concat([old_addr_apjc, df_apjc], sort=False, join='outer')

    # 提取org/bu数据
    for df in df_list:
        if df.columns.name != 'APJC':
            dfx = df.ADDRESSABLE.copy()
            # 以下将Series转置并与原使数据合并
            dic = {}
            for bu in dfx.index:
                dic[bu] = dfx.loc[bu]
            dfx = pd.DataFrame(dic, index=[0])
            dfx.loc[:, 'DATE'] = pd.Timestamp.now().strftime('%Y-%m-%d')
            dfx.set_index('DATE', inplace=True)

            if df.columns.name == 'FOC':
                df_foc = pd.concat([old_addr_foc, dfx], sort=False, join='outer')
            elif df.columns.name == 'FDO':
                df_fdo = pd.concat([old_addr_fdo, dfx], sort=False, join='outer')
            elif df.columns.name == 'SHK':
                df_shk = pd.concat([old_addr_shk, dfx], sort=False, join='outer')
            elif df.columns.name == 'JPE':
                df_jpe = pd.concat([old_addr_jpe, dfx], sort=False, join='outer')
            elif df.columns.name == 'NCB':
                df_ncb = pd.concat([old_addr_ncb, dfx], sort=False, join='outer')

    # put the 'DATE' out for duplicates removal
    df_apjc.reset_index(inplace=True)
    df_foc.reset_index(inplace=True)
    df_fdo.reset_index(inplace=True)
    df_shk.reset_index(inplace=True)
    df_jpe.reset_index(inplace=True)
    df_ncb.reset_index(inplace=True)

    # drop duplicates if date and contents are same
    df_apjc.drop_duplicates(keep='first', inplace=True)
    df_foc.drop_duplicates(keep='first', inplace=True)
    df_fdo.drop_duplicates(keep='first', inplace=True)
    df_shk.drop_duplicates(keep='first', inplace=True)
    df_jpe.drop_duplicates(keep='first', inplace=True)
    df_ncb.drop_duplicates(keep='first', inplace=True)

    df_apjc.set_index('DATE', inplace=True)
    df_foc.set_index('DATE', inplace=True)
    df_fdo.set_index('DATE', inplace=True)
    df_shk.set_index('DATE', inplace=True)
    df_jpe.set_index('DATE', inplace=True)
    df_ncb.set_index('DATE', inplace=True)

    df_dict = {'APJC': df_apjc,
               'FOC': df_foc,
               'FDO': df_fdo,
               'SHK': df_shk,
               'JPE': df_jpe,
               'NCB':df_ncb
               }
    del old_addr_apjc,old_addr_foc,old_addr_fdo,old_addr_shk,old_addr_jpe,old_addr_ncb
    gc.collect()

    return df_dict,df_apjc,df_foc,df_fdo,df_shk,df_jpe,df_ncb

def collect_new_addr_and_comb_historical_emea(df_list, addr_fname):
    '''
    This program convert the df_list (addressable backlog summary list) formats, mark it with current date for hte latest record. And concat it with what's already in the addressable_tracker.
    :param df_list: Summary created which is a list of df: addr_df_summary
    :param addr_fname: file name of the tracker file
    :return:
    '''
    old_addr_emea = pd.read_excel(addr_fname, sheet_name='EMEA', index_col='DATE')
    old_addr_fcz = pd.read_excel(addr_fname, sheet_name='FCZ', index_col='DATE')
    old_addr_fve = pd.read_excel(addr_fname, sheet_name='FVE', index_col='DATE')

    # 提取apjc by org数据
    df_emea = pd.concat(df_list, sort=False)
    df_emea = df_emea.loc[['EMEA', 'FCZ', 'FVE']]
    df_emea = df_emea.ADDRESSABLE

    # 以下将Series转置并于原使数据合并
    dic = {}
    for org in df_emea.index:
        dic[org] = df_emea.loc[org]
    df_emea = pd.DataFrame(dic, index=[pd.Timestamp.now().strftime('%Y-%m-%d')])
    df_emea.index.name = 'DATE'
    df_emea = pd.concat([old_addr_emea, df_emea], sort=False, join='outer')

    # 提取org/bu数据
    for df in df_list:
        if df.columns.name != 'EMEA':
            dfx = df.ADDRESSABLE
            # 以下将Series转置并与原使数据合并
            dic = {}
            for bu in dfx.index:
                dic[bu] = dfx.loc[bu]
            dfx = pd.DataFrame(dic, index=[0])
            dfx.loc[:, 'DATE'] = pd.Timestamp.now().strftime('%Y-%m-%d')
            dfx.set_index('DATE', inplace=True)

            if df.columns.name == 'FCZ':
                df_fcz = pd.concat([old_addr_fcz, dfx], sort=False, join='outer')
            elif df.columns.name == 'FVE':
                df_fve = pd.concat([old_addr_fve, dfx], sort=False, join='outer')

    # put the 'DATE' out for duplicates removal
    df_emea.reset_index(inplace=True)
    df_fcz.reset_index(inplace=True)
    df_fve.reset_index(inplace=True)

    df_emea.drop_duplicates(keep='first', inplace=True)
    df_fcz.drop_duplicates(keep='first', inplace=True)
    df_fve.drop_duplicates(keep='first', inplace=True)

    df_emea.set_index('DATE', inplace=True)
    df_fcz.set_index('DATE', inplace=True)
    df_fve.set_index('DATE', inplace=True)

    df_dict = {'EMEA': df_emea,
               'FCZ': df_fcz,
               'FVE': df_fve,
               }
    del old_addr_emea,old_addr_fcz,old_addr_fve
    gc.collect()

    return df_dict,df_emea,df_fcz,df_fve

def collect_new_addr_and_comb_historical_americas(df_list, addr_fname):
    '''
    This program convert the df_list (addressable backlog summary list) formats, mark it with current date for hte latest record. And concat it with what's already in the addressable_tracker.
    :param df_list: Summary created which is a list of df: addr_df_summary
    :param addr_fname: file name of the tracker file
    :return:
    '''
    old_addr_americas = pd.read_excel(addr_fname, sheet_name='Americas', index_col='DATE')
    old_addr_ftx = pd.read_excel(addr_fname, sheet_name='FTX', index_col='DATE')
    old_addr_tau = pd.read_excel(addr_fname, sheet_name='TAU', index_col='DATE')
    old_addr_sjz = pd.read_excel(addr_fname, sheet_name='SJZ', index_col='DATE')
    old_addr_jmx = pd.read_excel(addr_fname, sheet_name='JMX', index_col='DATE')
    old_addr_fgu = pd.read_excel(addr_fname, sheet_name='FGU', index_col='DATE')
    old_addr_fjz = pd.read_excel(addr_fname, sheet_name='FJZ', index_col='DATE')
    old_addr_tsp = pd.read_excel(addr_fname, sheet_name='TSP', index_col='DATE')

    # 提取apjc by org数据
    df_americas = pd.concat(df_list, sort=False)
    df_americas = df_americas.loc[['Americas', 'FTX', 'SJZ','TAU','JMX','FGU','FJZ','TSP']]
    df_americas = df_americas.ADDRESSABLE

    # 以下将Series转置并于原使数据合并
    dic = {}
    for org in df_americas.index:
        dic[org] = df_americas.loc[org]
    df_americas = pd.DataFrame(dic, index=[pd.Timestamp.now().strftime('%Y-%m-%d')])
    df_americas.index.name = 'DATE'
    df_americas = pd.concat([old_addr_americas, df_americas], sort=False, join='outer')

    # 提取org/bu数据
    for df in df_list:
        if df.columns.name != 'Americas':
            dfx = df.ADDRESSABLE
            # 以下将Series转置并与原使数据合并
            dic = {}
            for bu in dfx.index:
                dic[bu] = dfx.loc[bu]
            dfx = pd.DataFrame(dic, index=[0])
            dfx.loc[:, 'DATE'] = pd.Timestamp.now().strftime('%Y-%m-%d')
            dfx.set_index('DATE', inplace=True)

            if df.columns.name == 'FTX':
                df_ftx = pd.concat([old_addr_ftx, dfx], sort=False, join='outer')
            elif df.columns.name == 'TAU':
                df_tau = pd.concat([old_addr_tau, dfx], sort=False, join='outer')
            elif df.columns.name == 'SJZ':
                df_sjz = pd.concat([old_addr_sjz, dfx], sort=False, join='outer')
            elif df.columns.name == 'JMX':
                df_jmx = pd.concat([old_addr_jmx, dfx], sort=False, join='outer')
            elif df.columns.name == 'FGU':
                df_fgu = pd.concat([old_addr_fgu, dfx], sort=False, join='outer')
            elif df.columns.name == 'FJZ':
                df_fjz = pd.concat([old_addr_fjz, dfx], sort=False, join='outer')
            elif df.columns.name == 'TSP':
                df_tsp = pd.concat([old_addr_tsp, dfx], sort=False, join='outer')

    # put the 'DATE' out for duplicates removal
    df_americas.reset_index(inplace=True)
    df_ftx.reset_index(inplace=True)
    df_tau.reset_index(inplace=True)
    df_sjz.reset_index(inplace=True)
    df_fgu.reset_index(inplace=True)
    df_jmx.reset_index(inplace=True)
    df_fjz.reset_index(inplace=True)
    df_tsp.reset_index(inplace=True)

    df_americas.drop_duplicates(keep='first', inplace=True)
    df_ftx.drop_duplicates(keep='first', inplace=True)
    df_tau.drop_duplicates(keep='first', inplace=True)
    df_sjz.drop_duplicates(keep='first', inplace=True)
    df_fgu.drop_duplicates(keep='first', inplace=True)
    df_jmx.drop_duplicates(keep='first', inplace=True)
    df_fjz.drop_duplicates(keep='first', inplace=True)
    df_tsp.drop_duplicates(keep='first', inplace=True)

    df_americas.set_index('DATE', inplace=True)
    df_ftx.set_index('DATE', inplace=True)
    df_tau.set_index('DATE', inplace=True)
    df_sjz.set_index('DATE', inplace=True)
    df_fgu.set_index('DATE', inplace=True)
    df_jmx.set_index('DATE', inplace=True)
    df_fjz.set_index('DATE', inplace=True)
    df_tsp.set_index('DATE', inplace=True)

    df_dict = {'Americas': df_americas,
               'FTX': df_ftx,
               'TAU': df_tau,
               'SJZ':df_sjz,
               'FGU':df_fgu,
               'JMX':df_jmx,
               'FJZ':df_fjz,
               'TSP':df_tsp
               }
    del old_addr_americas,old_addr_ftx,old_addr_tau,old_addr_sjz,old_addr_fgu,old_addr_jmx,old_addr_fjz,old_addr_tsp
    gc.collect()

    return df_dict,df_americas,df_ftx,df_tau,df_sjz,df_fgu,df_jmx,df_fjz,df_tsp


def create_backlog_distribution_summary_by_bu(df, org, bu, backlog_distribution_date_labels, data_col='C_UNSTAGED_QTY'):
    '''
    针对一个确定的BU,按照Main_PF对backlog进行按周、按date_label汇总
    :param df: df_3a4
    :param org: ORG code
    :param bu: BU name (filter by BUSINESS_UNIT)
    :param backlog_distribution_date_labels: Date labels for the summaries
    :param data_col: col for data summaries
    :return:
    '''
    df_base = pd.DataFrame()

    for label in backlog_distribution_date_labels:
        df_label = pd.DataFrame()
        df_bu = df[(df.ORGANIZATION_CODE == org) & (df.BUSINESS_UNIT == bu)  & (df.MFG_HOLD != 'Y')].copy()
        df_bu = df_bu[df_bu[label].notnull()].copy()  # 去除空值，以免某一个pf全是空值时出错
        df_bu.set_index(label, inplace=True)

        pf_list = df_bu.main_pf.unique()
        for pf in pf_list:
            dfy = df_bu[df_bu.main_pf == pf].copy()
            dfy = dfy.resample('W-MON',label='left').sum()

            dfy = dfy[data_col]

            # 删除0值
            for l in dfy.index:
                if dfy[l]==0:
                    dfy.drop(l,axis=0,inplace=True)

            # 以下讲Series转置并于原使数据合并
            dic = {}
            for wk in dfy.index:
                dic[wk.strftime('%m-%d-%Y')] = dfy.loc[wk]
            df_pf = pd.DataFrame(dic, index=[pf])
            df_pf.index.name = label
            df_label = pd.concat([df_label, df_pf], sort=False, join='outer')

        df_label.loc[label + ':TOTAL', :] = df_label.sum(axis=0)
        df_label.loc[:, 'DATE_LABEL'] = label

        df_base = pd.concat([df_base, df_label], sort=False, join='outer')

    df_base.loc[:, 'ORGANIZATION_CODE'] = org
    df_base.loc[:, 'BUSINESS_UNIT'] = bu

    df_base.index.name = data_col
    df_base.reset_index(inplace=True)

    df_base.set_index(['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'DATE_LABEL', data_col], inplace=True)

    df_base.loc[:, 'Total'] = df_base.sum(axis=1)


    return df_base

def create_df_for_dfpm(df,dfpm,dfpm_mapping_to,col_3a4_dfpm):
    '''
    Create df based on dfpm and mapping
    :param df:
    :param dfpm:
    :param dfpm_mapping:
    :return:
    '''
    df_dfpm = pd.DataFrame()
    for org, content in dfpm_mapping_to[dfpm].items():
        dfx = df[(df.ORGANIZATION_CODE == org) & (df.main_bu.isin(content[0])) & (~df.PRODUCT_FAMILY.isin(content[2]))]
        dfy = df[(df.ORGANIZATION_CODE == org) & (df.PRODUCT_FAMILY.isin(content[1]))]

        df_dfpm = pd.concat([df_dfpm, dfx, dfy], join='outer', sort=False)

    df_dfpm=df_dfpm[col_3a4_dfpm]
    df_dfpm.set_index('ORGANIZATION_CODE',inplace=True)

    return df_dfpm

def create_backlog_distribution_summary_by_bu_new(df_dfpm, backlog_distribution_date_labels,
                                                  data_col='C_UNSTAGED_QTY',func='SUM',by='PO_NUMBER',excl_packed=True):
    '''
    针对一个确定的BU,按照Main_PF对backlog进行按周、按date_label汇总
    :param df: df_3a4
    :param org: ORG code
    :param bu: BU name (filter by BUSINESS_UNIT)
    :param backlog_distribution_date_labels: Date labels for the summaries
    :param data_col: col for data summaries
    :return:
    '''

    #去除df_dfpm重的minor pid及MFG hold
    df_dfpm_clean=df_dfpm[(df_dfpm.MFG_HOLD!='Y')].copy()
    # 按定义的条件处理
    if excl_packed:
        df_dfpm_clean=df_dfpm_clean[df_dfpm_clean.PACKOUT_QUANTITY!='Packout Completed'].copy()
    if by=='SO_SS': # count 数量时使用
        df_dfpm_clean.drop_duplicates('SO_SS',inplace=True)

    df_summary = pd.DataFrame()
    for label in backlog_distribution_date_labels:
        if func=='SUM':
            df_label=df_dfpm_clean.pivot_table(index=['ORGANIZATION_CODE','BUSINESS_UNIT','PRODUCT_FAMILY'],
                                         columns=label,
                                         values=data_col,
                                         aggfunc=sum)
        elif func=='COUNT':
            df_label = df_dfpm_clean.pivot_table(index=['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'PRODUCT_FAMILY'],
                                           columns=label,
                                           values=data_col,
                                           aggfunc=len)

        df_label.loc[:,'DATE_LABEL']=label
        df_label.reset_index(inplace=True)
        df_label.set_index(['DATE_LABEL','ORGANIZATION_CODE','BUSINESS_UNIT','PRODUCT_FAMILY'],inplace=True)
        #df_label.loc[('','','','BU subtotal'),:]=df_label.sum(axis=0)

        df_summary=pd.concat([df_summary,df_label],sort=False,join='outer')

    df_summary=df_summary.resample('W-MON',axis=1,label='left').sum()

    df_summary.columns=df_summary.columns.map(lambda x: x.strftime('%m-%d-%Y'))

    df_summary.loc[:, 'Total'] = df_summary.sum(axis=1)

    # 改变表头名称
    if func=='SUM' and by=='PO_NUMBER' and data_col=='C_UNSTAGED_QTY':
        df_summary.index.names=['Unstaged QTY(excl. packed&mfg hold)','','BUSINESS_UNIT','PRODUCT_FAMILY']
    elif func=='SUM' and by=='PO_NUMBER' and data_col=='po_unstaged_rev':
        df_summary.index.names = ['Unstaged M$(excl. packed&mfg hold)', '','BUSINESS_UNIT','PRODUCT_FAMILY']
        df_summary=df_summary.applymap(lambda x: round(x/1000000,2))
    elif func == 'COUNT' and by == 'PO_NUMBER' and data_col == 'PO_NUMBER':
        df_summary.index.names = ['Unstaged PO Count(excl. packed&mfg hold)', '','BUSINESS_UNIT','PRODUCT_FAMILY']
    elif func == 'COUNT' and by == 'SO_SS' and data_col == 'SO_SS':
        df_summary.index.names = ['Unstaged SS Count(excl. packed&mfg hold)', '','BUSINESS_UNIT','PRODUCT_FAMILY']

    df_summary.loc[('','','','total'),:]=df_summary.sum(axis=0)
    # 删除0 值col
    for col in df_summary.columns:
        if df_summary.loc[('','','','total'),col]==0:
            df_summary.drop(col,axis=1,inplace=True)
    df_summary.drop(('','','','total'),axis=0,inplace=True)

    return df_summary

def create_backlog_distribution_summary_by_org_new(df, backlog_distribution_date_labels,
                                                  data_col='C_UNSTAGED_QTY',func='SUM',by='PO_NUMBER',excl_packed=True):
    '''
    针对一个确定的BU,按照Main_PF对backlog进行按周、按date_label汇总
    :param df: df_3a4
    :param org: ORG code
    :param bu: BU name (filter by BUSINESS_UNIT)
    :param backlog_distribution_date_labels: Date labels for the summaries
    :param data_col: col for data summaries
    :return:
    '''


    #去除minor及MFG hold
    df_clean=df[(df.MFG_HOLD!='Y')].copy()
    # 按定义的条件处理
    if excl_packed:
        df_clean=df_clean[df_clean.PACKOUT_QUANTITY!='Packout Completed'].copy()

    if by=='SO_SS': # count 数量时使用
        df_clean.drop_duplicates('SO_SS',inplace=True)

    df_summary = pd.DataFrame()
    for label in backlog_distribution_date_labels:
        if func=='SUM':
            df_label=df_clean.pivot_table(index=['ORGANIZATION_CODE'],
                                         columns=label,
                                         values=data_col,
                                         aggfunc=sum)
        elif func=='COUNT':
            df_label = df_clean.pivot_table(index=['ORGANIZATION_CODE'],
                                           columns=label,
                                           values=data_col,
                                           aggfunc=len)

        df_label.loc[:,'DATE_LABEL']=label
        df_label.reset_index(inplace=True)
        df_label.set_index(['DATE_LABEL','ORGANIZATION_CODE'],inplace=True)
        #df_label.loc[('','','','BU subtotal'),:]=df_label.sum(axis=0)

        df_summary=pd.concat([df_summary,df_label],sort=False,join='outer')

    df_summary=df_summary.resample('W-MON',axis=1,label='left').sum()

    df_summary.columns=df_summary.columns.map(lambda x: x.strftime('%m-%d-%Y'))

    df_summary.loc[:, 'Total'] = df_summary.sum(axis=1)

    # 改变表头名称
    if func=='SUM' and by=='PO_NUMBER' and data_col=='C_UNSTAGED_QTY':
        df_summary.index.names=['Unstaged QTY(excl. packed&mfg hold)','']
    elif func=='SUM' and by=='PO_NUMBER' and data_col=='po_unstaged_rev':
        df_summary.index.names = ['Unstaged M$(excl. packed&mfg hold)', '']
        df_summary=df_summary.applymap(lambda x: round(x/1000000,2))
    elif func == 'COUNT' and by == 'PO_NUMBER' and data_col == 'PO_NUMBER':
        df_summary.index.names = ['Unstaged PO Count(excl. packed&mfg hold)', '']
    elif func == 'COUNT' and by == 'SO_SS' and data_col == 'SO_SS':
        df_summary.index.names = ['Unstaged SS Count(excl. packed&mfg hold)', '']

    # 删除0 值col
    df_summary.loc[('', 'total'), :] = df_summary.sum(axis=0)
    for col in df_summary.columns:
        if df_summary.loc[('', 'total'), col] == 0:
            df_summary.drop(col, axis=1, inplace=True)
    df_summary.drop(('', 'total'), axis=0, inplace=True)

    return df_summary

def create_backlog_distribution_summary_by_org(df, backlog_distribution_date_labels, data_col='C_UNSTAGED_QTY'):
    '''

    :param df:
    :param org_list:
    :param backlog_distribution_date_labels:
    :param data_col:
    :return:
    '''
    df_base = pd.DataFrame()

    for label in backlog_distribution_date_labels:
        dfx = df[ (df.MFG_HOLD != 'Y')& (df.PACKOUT_QUANTITY != 'PACKOUT_COMPLETED')]

        dfp=dfx.pivot_table(index='ORGANIZATION_CODE',columns=label,values=data_col,aggfunc=sum)
        dfp.resample('W-MON',axis=1,label='left').sum()
        dfp.columns=dfp.columns.map(lambda x: x.strftime('%m-%d-%Y'))

        dfp.loc[label + ':Total',:] = dfp.sum(axis=0)

        dfp.loc[:,'DATE_LABEL']=label
        df_base=pd.concat([df_base,dfp],sort=False,join='outer')

    df_base.reset_index(inplace=True)
    df_base.set_index(['DATE_LABEL', 'ORGANIZATION_CODE'], inplace=True)

    df_base.loc[:, 'Total'] = df_base.sum(axis=1)

    return df_base

def create_backlog_distribution_summary_by_date_org(df,frequency='WK',column_date='ORIGINAL_FCD_NBD_DATE',value='PO_NUMBER'):
    '''
    Create summary by original NBD for orders without mfg hold and is not fully packed
    :param df:
    :return:
    '''
    dfx = df[(df.PACKOUT_QUANTITY != 'Packout Completed')].copy()
    dfx.loc[:,'With_Hold']=np.where(dfx.ORDER_HOLDS.notnull(),'Yes','No')
    #TODO: change to use MFG_HOLD later if it looks all right

    if value=='PO_NUMBER':
        dfp=dfx.pivot_table(index=['ORGANIZATION_CODE','With_Hold'],columns=column_date,values=value,aggfunc=len).copy()
        dfp.index.names=['No. of PO','With_Hold']
    elif value=='SO_SS':
        dfx.drop_duplicates('SO_SS',inplace=True)
        dfp = dfx.pivot_table(index=['ORGANIZATION_CODE', 'With_Hold'], columns=column_date,
                             values=value, aggfunc=len).copy()
        dfp.index.names=['No. of SS','With_Hold']

    if frequency=='WK': # else: DAY
        dfp=dfp.resample('W-MON',axis=1,label='left').sum()

    dfp.columns = dfp.columns.map(lambda x: x.strftime('%m-%d-%Y'))

    dfp.loc[:, 'Total'] = dfp.sum(axis=1)
    dfp.loc[('All ORG','Total'),:] = dfp.sum(axis=0)

    # 删除0 值col
    for col in dfp.columns:
        if dfp.loc[('All ORG','Total'),col]==0:
            dfp.drop(col,axis=1,inplace=True)


    return dfp

def create_addr_charts(df_list, item_list, fname, region, org=None):
    '''
    Base on the summary df to create charts
    :param org: if creating by BU charts, then should specify org; no need to specify BU in this case
    :param df_list: Summary created which is a list of df: addr_df_summary
    :param item_list: either list of Org, or list of BU (redefine in code below)
    :param fname: chart address and name
    :return: None
    '''
    if org == None: # regional charts
        df = pd.concat(df_list, sort=False)
    else:  # BU chart for a specific org
        #print(df_list)
        for dfx in df_list:
            if dfx.columns.name == org:
                df = dfx.iloc[:-1].copy()
                df.sort_values(by='ADDRESSABLE', ascending=False, inplace=True)
                item_list = dfx.index.tolist()

                break

    df.reset_index(inplace=True)
    df = df[df.BU.isin(item_list)].copy()

    df.set_index('BU', inplace=True)
    df.index.name = ''

    # 不存在的元素用0补上
    addr_col=['ADDRESSABLE', 'NOT_ADDRESSABLE','UNSCHEDULED', 'MFG_HOLD',]
    not_existing_col=np.setdiff1d(addr_col,df.columns)
    for x in not_existing_col:
        df.loc[:,x]=0

    df = df[addr_col]
    #print(df)
    df.loc[:,'total']=df.sum(axis=1)
    df=df[df.total>0]
    df.drop('total',axis=1,inplace=True)
    # 只针对前17个BU制图
    df=df.iloc[:17,:]
    df.fillna(0, inplace=True)

    fig, ax = plt.subplots(1, 1, figsize=(8, 2))
    # plt.figure(figsize=(8,1))
    # plt.bar(df.index,df.Addressable,color='blue')
    # plt.bar(df.index, df.Not_addressable, color='orange',bottom=df.Addressable)

    if df.shape[0] == 0:   # if there is no data for the org use below dummy chart
        df_dummy = pd.Series(range(0, 800))
        ax.text(300, 4, 'No backlog!', color='white')
        df_dummy.plot(ax=ax, kind='hist', bins=100, rot=0, fontsize=5)
    else:
        df.plot(ax=ax, kind='bar', stacked=True, rot=0, fontsize=5)

    ax.set_ylabel('(M$)', fontsize=5)
    if df.columns.name == None:
        title = region + ' Backlog'
        value_size=4
    else:
        title = df.columns.name + ' Backlog'
        value_size=5

    ax.set_title(title, fontsize=7)
    ax.legend(df.columns, fontsize=5)

    #ax.spines['top'].set_visible(False)
    #ax.spines['right'].set_visible(False)
    #ax.spines['left'].set_visible(False)
    #print(df)
    for x in range(len(df.index)):
        value_1 = df.iloc[x, 0]
        value_2 = df.iloc[x, 1]
        value_3 = df.iloc[x, 2]
        value_4 = df.iloc[x, 3]

        # only write the value when>1
        if value_1 > 1:
            y=value_1/2
            ax.text(x, y,int(value_1),  fontsize=5)
        if value_2 > 1:
            y=value_1 + value_2 * 2/5
            ax.text(x, y, int(value_2), fontsize=value_size)
        if value_3 > 1:
            y=value_1 + value_2 + (value_3 * 2/5)
            ax.text(x, y, int(value_3), fontsize=value_size)
        if value_4 > 1:
            y=value_1 + value_2 + value_3 + (value_4 * 2/5)
            ax.text(x, y, int(value_4), fontsize=value_size)

    # max_y = int(df.Addressable.max()+df.Not_addressable.max())
    # y_tick = [x for x in range(0, max_y + 10, 10)]
    # ax.set_yticks(y_tick)
    # plt.yticklabels(y_tick,fontsize=5)
    # ax.grid(color='grey', linestyle='-', linewidth=1, alpha=0.2)

    # plt.show()
    plt.savefig(fname, bbox_inches='tight', dpi=200)

    del df,ax
    gc.collect()

def create_blg_charts(df_list, fname):
    '''
    Base on the summary df to create charts
    #:param org: if creating by BU charts, then should specify org; no need to specify BU in this case
    :param df_list: Summary created which is a list of df
    :param item_list: either list of Org, or list of BU
    :param fname: chart address and name
    :return: df
    '''

    for df in df_list:
        df.loc[:, 'APJC'] = df.sum(axis=1)
        df = df[df.APJC > 0].copy()
        df.drop('APJC', axis=1, inplace=True)
        df.fillna(0, inplace=True)
        # print(df)
        # can't control the legend fontsize and position, so not using pandas plot
        fig, axes = plt.subplots(len(df.columns), 1, sharey=True, sharex=True, figsize=(3, 1.5 * len(df.columns)))
        for col_index, col, ax, color in zip(range(len(df.columns)), df.columns, axes.ravel(),
                                             ['c', 'g', (0.2, 0.3, 0.5)]):
            ax.bar(df.index, df.iloc[:, col_index], color=color, label=col)

            if ax == axes.ravel()[0]:
                ax.set_title(df.columns.name, fontsize=7)
            ax.set_ylabel('PO Count', fontsize=3)

            for x in range(len(df.index)):
                y_value = int(df.iloc[x, col_index])
                # print(y_value)
                if y_value > 2:
                    ax.text(x, y_value, y_value, fontsize=5)

            ax.tick_params(labelsize=4)
            ax.set_xticklabels(df.index, rotation=45)

            ax.legend(loc='upper left', fontsize=4)
            # max_y = int(df.max().max())
            # y_tick = [x for x in range(0, max_y + 10, 10)]
            # print(y_tick)
            # ax.set_yticks(y_tick)
            # ax.set_yticklabels(y_tick, fontsize=5) # have to work together with set_yticks
            # ymajorLocator = MultipleLocator(10)
            # ax.yaxis.set_major_locator(ymajorLocator)

            # ax.grid(color='grey', linestyle='-', linewidth=1, alpha=0.2)

        plt.subplots_adjust(hspace=0)
        # plt.show()
        f_name = fname[:-4] + '_' + df.columns.name + '.png'
        plt.savefig(f_name, bbox_inches='tight')  # ,dpi=200)

        del df,fig,axes
        gc.collect()

    # return df



def read_compliance_from_smartsheet(df_3a4):
    '''
    Read CTB data from smartsheet - pick the latest record by org
    :param df_3a4: read ship_to_country from df_3a4
    :return: ctb df
    '''
    # 数据源基本设定 - smartsheet设定
    token = os.getenv('SMARTSHEET_TOKEN_WNBU_COMPLIANCE')
    sheet_id = os.getenv('WNBU_COMPLIANCE_SHEET_ID')
    # token='nwy98xheoeq6ns4ju33s34yvrq'
    # sheet_id='313000021256068'
    #print('compliance token:',token)

    proxies = None  # for proxy server

    compliance_error_msg = []
    compliance_read_msg = []

    # 读取smartsheet的对象（从smartsheet_hanndler导入类）
    smartsheet_client = SmartSheetClient(token, proxies)

    # 从smartsheet读取compliance data
    df = smartsheet_client.get_sheet_as_df(sheet_id, add_row_id=False, add_att_id=False)

    df.dropna(how='all', inplace=True)

    df.set_index('Domain', inplace=True)

    domain = df.index.values
    pid = df.columns.values

    all_country = list(df_3a4.SHIP_TO_COUNTRY.unique())

    no_ship = {}
    for p in pid:
        for d in domain:
            rule = df.loc[d, p]
            if rule == 'EVERY': # all ready, emply restriction list
                no_ship[p + '-' + d] = []
            elif rule == 'DONT SHIP' or rule == None:  # All Not ready; or 如果table有空值
                no_ship[p + '-' + d] = all_country
            else: # Not ready country
                rule = rule.split('\n')
                rule = [x.strip() for x in rule]
                no_ship[p + '-' + d] = rule

    df.reset_index(inplace=True)

    return df, no_ship


def check_str(x, y_list):
    '''
    Check if items in y_list exists in string x
    :param x: string
    :param y_list: list of string
    :return: return str if str exist in string x
    '''

    for str in y_list:
        if str in x:
            return str


def check_xy_in_z(x, y, z):
    if z != ' ':
        if x in z or y in z:
            return 'keep_hold'
        else:
            return 'ok_release'


def check_compliance_for_wnbu(df_3a4, no_ship):
    """
    Check for WNBU if compliance hold can be released
    """
    dfx=df_3a4[df_3a4.main_bu=='WNBU'].copy()

    # check the pid is matching with the no_ship smartsheet table: if compliance hold and PID maps
    dfx.loc[:, 'wnbu_compliance'] = np.where(dfx.ORDER_HOLDS.notnull(),
                                                np.where(dfx.ORDER_HOLDS.str.contains('Compliance',case=False),
                                                            dfx.PRODUCT_ID.map(lambda x: check_str(x, no_ship.keys())),
                                                            ' '),
                                                ' ')

    dfx.loc[:, 'wnbu_compliance'] = dfx.wnbu_compliance.map(
        lambda x: no_ship.get(x) if x in no_ship.keys() else ' ')  # has to be ' ' otherewise below check_xy_in_z would error

    # df_wnbu=df_3a4[(df_3a4.BUSINESS_UNIT=='WNBU') & (df_3a4.SHIP_TO_COUNTRY.notnull()) & (df_3a4.wnbu_compliance.notnull())].copy()

    dfx.loc[:, 'SHIP_TO_COUNTRY'].fillna('', inplace=True)
    dfx.loc[:, 'wnbu_compliance'] = dfx.apply(
        lambda x: check_xy_in_z(x.SHIP_TO_COUNTRY, x.END_TO_COUNTRY, x.wnbu_compliance),
        axis=1)

    df_compliance = dfx[(dfx.wnbu_compliance == 'keep_hold') | (dfx.wnbu_compliance == 'ok_release')].copy()

    df_compliance.loc[:, 'wnbu_compliance'] = np.where(df_compliance.SHIP_TO_COUNTRY.isnull(),
                                                       'country_missing',
                                                       np.where(df_compliance.END_TO_COUNTRY.isnull(),
                                                                'country_missing',
                                                                df_compliance.wnbu_compliance)
                                                       )

    col = ['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'PO_NUMBER', 'PRODUCT_ID', 'LINE_CREATION_DATE','SHIP_TO_COUNTRY', 'END_TO_COUNTRY',
           'ORDER_HOLDS', 'wnbu_compliance']
    df_compliance = df_compliance[col]
    df_compliance_release = df_compliance[df_compliance.wnbu_compliance == 'ok_release'].copy()
    df_compliance_hold = df_compliance[df_compliance.wnbu_compliance == 'keep_hold'].copy()
    df_country_missing = df_compliance[df_compliance.wnbu_compliance == 'country_missing'].copy()

    del df_compliance,dfx
    gc.collect()

    return df_compliance_release, df_compliance_hold, df_country_missing


def read_ctb_from_smartsheet():
    '''
    Read CTB data from smartsheet - pick the latest record by org
    :return: ctb df
    '''
    # 数据源基本设定 - smartsheet设定
    token = os.getenv('SMARTSHEET_TOKEN_CTB')
    attachment_sheet_id = os.getenv('CTB_SHEET_ID')
    #print(token)
    #print(attachment_sheet_id)

    proxies = None  # for proxy server

    ctb_error_msg = []

    # 读取smartsheet的对象（从smartsheet_hanndler导入类）
    smartsheet_client = SmartSheetClient(token, proxies)

    # 从smartsheet读取attachment
    attachment_sheet_df = smartsheet_client.get_sheet_as_df(attachment_sheet_id, add_row_id=True, add_att_id=True)
    # 按照CM org保留最后的记录

    attachment_sheet_df.drop_duplicates(['CM'], keep='last', inplace=True)
    attachment_sheet_df.reset_index(inplace=True)
    attachment_sheet_df.drop('index', axis=1, inplace=True)

    # 将相应的attachment内容读入att_df并在smartsheet中做相应标识
    att_df = pd.DataFrame(columns=['SO_SS_LN', 'BUILD_DATE', 'CTB_STATUS', 'CTB_COMMENT'])
    for row in range(attachment_sheet_df.shape[0]):
        attachment_id = attachment_sheet_df.loc[row, 'attachment_id']
        row_id = attachment_sheet_df.loc[row, 'row_id']

        # print(row_id,attachment_id)
        # 读取附加内容
        att_df_new = smartsheet_client.get_attachment_per_row_as_df(attachment_id=attachment_id,
                                                                    sheet_id=attachment_sheet_id,
                                                                    row_id=row_id)

        # 对附件内容进行验证 - 格式正确则读取内容
        temp_col = ['SO_SS_LN', 'BUILD_DATE', 'CTB_STATUS', 'CTB_COMMENT']

        file_uploaded_by = attachment_sheet_df.loc[row, 'UPLOADED_BY']
        file_org = attachment_sheet_df.loc[row, 'CM']
        file_upload_date = attachment_sheet_df.loc[row, 'Created']

        missing_col = np.setdiff1d(temp_col, att_df_new.columns.values)

        if len(missing_col) > 0:
            update_dict = [{'STATUS': 'FORMAT_ERROR'}]
            # error_format_org=att_df_new.ORGANIZATION_CODE.unique() # not using this as org col may be missing

            msg = 'Latest CTB format error: {} file loaded by {} on {}'.format(
                file_org, file_uploaded_by, file_upload_date)

            ctb_error_msg.append(msg)
        else:
            if att_df_new.shape[0] > 0:
                read_date = pd.Timestamp.now().strftime('%Y-%m-%d')
                update_dict = [{'STATUS': 'COLLECTED', 'READ_DATE': read_date}]
                att_df = pd.concat([att_df, att_df_new], join='outer', sort=False)
                msg = 'CTB file used: {} file loaded by {} on {}'.format(
                    file_org, file_uploaded_by, file_upload_date)
            else:
                update_dict = [{'STATUS': 'EMPTY_CONTENT'}]
                msg = 'Latest CTB content empty: {} file loaded by {} on {}'.format(
                    file_org, file_uploaded_by, file_upload_date)

                ctb_error_msg.append(msg)

        # 更新smartsheet
        smartsheet_client.update_row_with_dict(ss=smartsheet.Smartsheet(token), process_type='update',
                                               sheet_id=attachment_sheet_id,
                                               row_id=int(attachment_sheet_df.iloc[row]['row_id']),
                                               update_dict=update_dict)
        if len(ctb_error_msg)>0:
            print('CTB error: \n', ctb_error_msg)

    att_df.rename(columns={'BUILD_DATE':'CM_CTB'},inplace=True)

    del attachment_sheet_df,att_df_new
    gc.collect()

    return att_df, ctb_error_msg



def create_addressable_summary_with_ctb(df_3a4):
    '''
    Create the summary of addressable with CTB info based on 3A4
    :param df_3a4:
    :return:
    '''
    if 'main_bu' in df_3a4.columns:
        index_org_bu=['ORGANIZATION_CODE', 'main_bu']
        index_org_bu_pf=['ORGANIZATION_CODE', 'main_bu', 'main_pf']
    else:
        index_org_bu = ['ORGANIZATION_CODE', 'BUSINESS_UNIT']
        index_org_bu_pf = ['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'PRODUCT_FAMILY']

    addressable_by_org_bu = df_3a4.pivot_table(index=index_org_bu,
                                               columns='ADDRESSABLE_FLAG', values='C_UNSTAGED_DOLLARS', aggfunc=sum)

    ctb_by_org_bu = df_3a4.pivot_table(index=index_org_bu,
                                       columns='CTB_STATUS', values='C_UNSTAGED_DOLLARS', aggfunc=sum)
    # Merge ctb to 3a4 by BU
    addr_ctb_by_org_bu = pd.merge(addressable_by_org_bu, ctb_by_org_bu, left_index=True, right_index=True, how='left')

    addressable_by_org_bu_pf = df_3a4.pivot_table(index=index_org_bu_pf,
                                                  columns='ADDRESSABLE_FLAG', values='C_UNSTAGED_DOLLARS', aggfunc=sum)

    ctb_by_org_bu_pf = df_3a4.pivot_table(index=index_org_bu_pf,
                                          columns='CTB_STATUS', values='C_UNSTAGED_DOLLARS', aggfunc=sum)
    # Merge ctb to 3a4 by BU and PF
    addr_ctb_by_org_bu_pf = pd.merge(addressable_by_org_bu_pf, ctb_by_org_bu_pf, left_index=True, right_index=True,
                                     how='left')

    addr_ctb_by_org_bu = addr_ctb_by_org_bu.applymap(lambda x: round(x / 1000000, 2))  # if x.notnull() else x)
    addr_ctb_by_org_bu_pf = addr_ctb_by_org_bu_pf.applymap(lambda x: round(x / 1000000, 2))  # if x.notnull() else x)
    """
    # Add subtotal by org for both pivot table
    df_w_total = []
    for org in org_name['APJC']:
        try:
            dfx = addr_ctb_by_org_bu.loc[(org, slice(None)), :].copy()
            dfx.loc[(org, org + ' Total'), :] = dfx.sum(axis=0)
            df_w_total.append(dfx)
        except:
            print('{} data not in 3a4, skip this org')


    addr_ctb_by_org_bu = pd.concat(df_w_total, sort=False)

    # print(addr_ctb_by_org_bu_pf)

    df_w_total = []
    for org in org_name['APJC']:
        dfx = addr_ctb_by_org_bu_pf.loc[(org, slice(None)), :].copy()
        dfx.loc[(org, 'Sub total', org + ' Total'), :] = dfx.sum(axis=0)
        # print(dfx)

        df_w_total.append(dfx)

    addr_ctb_by_org_bu_pf = pd.concat(df_w_total, sort=False)
    
    """

    return addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf


def create_ctb_summary_for_material_gating(df_3a4):
    '''
    Create a subset of CTB summary for Material gating lines only
    :param df_3a4:
    :return:
    '''

    df_ctb_material = df_3a4[df_3a4.CTB_STATUS.isin(['MATERIAL', 'Material', 'material'])].copy()
    # ctb_summary_for_material = df_ctb_material.pivot_table(index=['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'CTB_COMMENT','main_pf'], aggfunc=sum)
    # ctb_summary_for_material=ctb_summary_for_material[['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'CTB_COMMENT','main_pf','C_UNSTAGED_DOLLARS', 'C_UNSTAGED_QTY']]
    if 'main_bu' in df_ctb_material.columns:
        ctb_summary_for_material = df_ctb_material[
            ['ORGANIZATION_CODE', 'main_bu', 'CTB_COMMENT', 'main_pf', 'C_UNSTAGED_DOLLARS', 'C_UNSTAGED_QTY']].copy()
    else:
        ctb_summary_for_material = df_ctb_material[
            ['ORGANIZATION_CODE', 'BUSINESS_UNIT', 'CTB_COMMENT', 'PRODUCT_FAMILY', 'C_UNSTAGED_DOLLARS', 'C_UNSTAGED_QTY']].copy()

    ctb_summary_for_material.sort_values(by=['ORGANIZATION_CODE', 'C_UNSTAGED_DOLLARS'], ascending=False, inplace=True)
    ctb_summary_for_material.set_index('ORGANIZATION_CODE', inplace=True)

    ctb_summary_for_material.C_UNSTAGED_DOLLARS = ctb_summary_for_material.C_UNSTAGED_DOLLARS.map(
        lambda x: round(x / 1000000, 2))

    del df_ctb_material
    gc.collect()

    return ctb_summary_for_material


def write_excel_file(fname, data_to_write):
    '''
    Write the df into excel files as different sheets
    :param fname: fname of the output excel
    :param df_to_write: a dict that contains {sheet_name:df}
    :return: None
    '''
    # engine='xlsxwriter' is used to avoid illegal character which lead to failure of saving the file
    writer = pd.ExcelWriter(fname,engine='xlsxwriter')

    for sheet_name, df in data_to_write.items():
        df.to_excel(writer, sheet_name=sheet_name)

    writer.save()



def create_outlier_chart(df_3a4, org, outlier_elements, fname_list):
    '''
    Create outlier chart - hist chart
    :param df_3a4:
    :param org: APJC, or each org name
    :param outlier_elements: Outlier columns: (df,threshold) mapping - for making chart and df summary
    :param outlier_thres_holds: Deccide to make chart only for days base on this threshold
    :param fname: output name of the chart
    :return:
    '''
    if org.upper() == 'APJC':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['FOC', 'FDO', 'SHK', 'JPE', 'NCB'])].copy()
    elif org.upper() == 'FOC':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['FOC'])].copy()
    elif org.upper() == 'FDO':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['FDO'])].copy()
    elif org.upper() == 'SHK':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['SHK'])].copy()
    elif org.upper() == 'JPE':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['JPE'])].copy()
    elif org.upper() == 'NCB':
        dfx = df_3a4[df_3a4.ORGANIZATION_CODE.isin(['NCB'])].copy()

    for element, thresh_hold, fname in zip(outlier_elements.keys(), [x[0] for x in outlier_elements.values()],
                                           fname_list):
        fig, ax = plt.subplots(1, 1, figsize=(8, 2))
        if element == 'ss_partial_staged_days':
            dfx = dfx.drop_duplicates('SO_SS', keep='first')
            dfx = dfx[dfx.ADDRESSABLE_FLAG != 'PO_CANCELLED']
            ax.set_ylabel('No. of SS', fontsize=5)
            ax.set_title(element.upper() + ">" + str(thresh_hold) + ' days (excl. cancelled)', fontsize=7)
        else:
            dfx = dfx.drop_duplicates('PO_NUMBER', keep='first')
            ax.set_ylabel('No. of PO', fontsize=5)
            ax.set_title(element.upper() + ">" + str(thresh_hold) + ' days', fontsize=7)

        ''' Below for making same ylim charts for comparison purpose
        if element=='entered_not_booked':
            ax.set_ylim([0,250])
            #ax.set_xlim([20,550])
        elif element=='booked_not_schedule':
            ax.set_ylim([0,650])
            #ax.set_xlim([20,700])
        elif element=='missed_ossd':
            ax.set_ylim([0,10])
            #ax.set_xlim([100,500])
        elif element=='cancel_aging_days':
            ax.set_ylim([0,125])
            #ax.set_xlim([20,1200])
        elif element=='ss_partial_staged_days':
            ax.set_ylim([0,25])
            #ax.set_xlim([50,150])
        '''

        # if dfx is empty, make the chart different
        plot_data = dfx[dfx[element] > thresh_hold][element]
        if len(plot_data) == 0:
            plot_data = pd.Series(range(0, 1000))
            ax.text(300, 5, 'Congratulations, no outlier!!', color='white')

        plot_data.plot(ax=ax, kind='hist', bins=100, rot=0, fontsize=5)

        # 3a4.hist(df_3a4[element][df_3a4[element] > thresh_hold], 100)

        # ax.set_xticks(range(1,int(df[data].max())),50)
        # ax.set_xticklabels([str(x) for x in range(1,int(df[data].max()),10)])
        # ax.set_ylim(0,max_y)

        ax.set_xlabel('Days', fontsize=5)

        plt.savefig(fname, bbox_inches='tight', dpi=200)

    del dfx,fig,ax
    gc.collect()


def create_outlier_df(df_3a4, outlier_elements):
    '''
    Create the outlier df by elements (for render the template) and save file to disk
    :param df_3a4:
    :param outlier_elements:
    :return:
    '''

    elements = [x for x in outlier_elements.keys()]
    threshold = [x[0] for x in outlier_elements.values()]

    df_not_booked = df_3a4[(df_3a4[elements[0]] > threshold[0])].copy()
    df_not_scheduled = df_3a4[(df_3a4[elements[1]] > threshold[1])].copy()
    df_not_packed = df_3a4[(df_3a4[elements[2]] > threshold[2])].copy()
    df_aging_cancel = df_3a4[(df_3a4[elements[3]] > threshold[3])].copy()
    df_partial_staged = df_3a4[(df_3a4[elements[4]] > threshold[4])].copy()
    df_missed_recommit=df_3a4[(df_3a4[elements[5]] > threshold[5])].copy()

    # print(df_partial_staged)
    # partial staged is based on SS - take below extra steps
    df_partial_staged.drop_duplicates('SO_SS', keep='first', inplace=True)
    df_partial_staged = df_partial_staged[df_partial_staged.ADDRESSABLE_FLAG != 'PO_CANCELLED'].copy()

    # Sort by aging days
    df_not_booked.sort_values(by=elements[0], ascending=False, inplace=True)
    df_not_scheduled.sort_values(by=elements[1], ascending=False, inplace=True)
    df_not_packed.sort_values(by=elements[2], ascending=False, inplace=True)
    df_aging_cancel.sort_values(by=elements[3], ascending=False, inplace=True)
    df_partial_staged.sort_values(by=elements[4], ascending=False, inplace=True)
    df_missed_recommit.sort_values(by=elements[5], ascending=False, inplace=True)

    # 以下确保关键列被按顺序提取排列到前面
    df_not_booked = df_not_booked.set_index(col_outlier['df_not_booked']).reset_index()
    df_not_scheduled = df_not_scheduled.set_index(col_outlier['df_not_scheduled']).reset_index()
    df_not_packed = df_not_packed.set_index(col_outlier['df_not_packed']).reset_index()
    df_aging_cancel = df_aging_cancel.set_index(col_outlier['df_aging_cancel']).reset_index()
    df_partial_staged = df_partial_staged.set_index(col_outlier['df_partial_staged']).reset_index()
    df_missed_recommit = df_missed_recommit.set_index(col_outlier['df_missed_recommit']).reset_index()

    # df_not_scheduled做进一步的处理
    df_not_scheduled=df_not_scheduled[df_not_scheduled.ORDER_HOLDS.isnull()].copy()

    # For Partial staged 去除PO_number col
    df_partial_staged.drop('PO_NUMBER', axis=1, inplace=True)

    # print(df_partial_staged)

    return df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel, df_partial_staged, df_missed_recommit


def create_outlier_spreadsheet(full_path, df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel,
                               df_partial_staged,df_missed_recommit):
    '''
    This create the outlier spreadsheet based on the whole dataset
    :param full_path:
    :param df_not_booked:
    :param df_not_scheduled:
    :param df_not_packed:
    :param df_aging_cancel:
    :param df_partial_staged:
    :return:
    '''

    #print(df_not_booked.head())
    #print(df_not_booked.columns)

    df_not_booked.set_index('ORGANIZATION_CODE', inplace=True)
    df_not_scheduled.set_index('ORGANIZATION_CODE', inplace=True)
    df_not_packed.set_index('ORGANIZATION_CODE', inplace=True)
    df_aging_cancel.set_index('ORGANIZATION_CODE', inplace=True)
    df_partial_staged.set_index('ORGANIZATION_CODE', inplace=True)
    df_missed_recommit.set_index('ORGANIZATION_CODE',inplace=True)

    df_dict = {'entered_not_booked': df_not_booked,
               'booked_not_scheduled': df_not_scheduled,
               'scheduled_not_packed': df_not_packed,
               'partial_staged': df_partial_staged,
               'aging_cancel': df_aging_cancel,
               'missed_recommit':df_missed_recommit,
               }

    # Write multiple DF to excel based on df_dict
    write_excel_file(full_path, df_dict)


def create_addr_trending_chart(df_dict,backlog_chart):
    '''
    Create addressable trending line chart based on addressable tracker
    :param df_dict: a dict by org from addressable tracker
    :return:
    '''

    for org, df in df_dict.items():
        #df.index = df.index.strftime('%m-%d')
        # 针对每个org (非APJC),取前6大的BU（根据过去20天的平均值）
        if org != 'APJC' and org!='EMEA' and org !='Americas':
            df = df.iloc[:, 1:].copy()
            df.loc['AVERAGE'] = df[-20:].mean()

            col = df.columns
            val = df.loc['AVERAGE'].values
            # 排序，选取前6大BU
            dfx = pd.DataFrame(index=col, data=val, columns=['Average'])
            dfx.sort_values(by='Average', inplace=True, ascending=False)
            col = dfx.index[:8]
            # 去掉Average行
            df = df[col][:-1].copy()

            #print(df)

        fig, ax = plt.subplots(1, 1, figsize=(8, 2))
        df.plot(ax=ax, kind='line', rot=0, fontsize=5, grid=True, linewidth=0.9)
        ax.grid(linestyle='-', linewidth=0.2)
        #print(df)
        #print(df.index)
        ax.set_ylabel('(M$)', fontsize=5)
        if org=='APJC' or org=='EMEA' or org=='Americas':
            ax.set_title(org + ' Addressable trending', fontsize=7)
        else:
            ax.set_title(org + ' Addressable trending (top BU)', fontsize=7)

        # ax.set_xticks(range(len(df.index)))
        # ax.set_xticklabels(df.index.strftime('%m-%d'),fontsize=5)
        ax.set_xlabel(None)

        ax.legend(df.columns, fontsize=4,loc='upper left')
        #plt.show()
        fname=backlog_chart[org.lower() + '_add_trending']
        plt.savefig(fname, bbox_inches='tight', dpi=200)

    del dfx,ax
    gc.collect()


def select_col(df_3a4, col_3a4):
    '''
    选取需要的columns
    '''

    df_3a4 = df_3a4[col_3a4]

    return df_3a4


def create_dic(df):
    '''
    生成config字典来记录根据df（pivoted)中的unique config
    :param df: pivoted df with PO as index, PIDs as columns
    :return: dictionary （PO作为键,Pid:order qty组合作为值）
    '''

    ref_dic = {}

    for i in range(df.shape[0]):
        df_line = df.iloc[i].dropna()

        sub_dic = {}
        for pid, qty in zip(df_line.index, df_line.values):
            sub_dic[pid] = qty

        ref_dic[df_line.name] = sub_dic

    return ref_dic


def compare_input_and_ref_dic(blg_dic, rtv_dic):
    '''
    比较两个config dic，根据相同的config得到比较结果
    :param blg_dic: dic to match with the base dic
    :param rtv_dic: config dic used as the base
    :return: val_dic (key: blg_dic keys, value: [rtv_dic keys])
    '''

    val_dic = {}
    for key, value in blg_dic.items():
        matches = filter(lambda x: x[1] == value, rtv_dic.items())

        # 存储相同config的订单列表
        order_list = []
        for order, _ in matches:
            order_list.append(order)

        val_dic[key] = order_list

    return val_dic


def create_config_matching_rtv_col(df, org_list):
    '''
    Create a col to indicate same config orders as the RTV orders
    :param df: df (unpivoted df)
    :param org_list: list of org code that will take this action - this needs enhancement to avoid mapping to other org's RTV order
    :return: df_3a4
    '''
    # split out the df based on waiting RTV and not cancelled PO
    df_rtv = df[(df.ADDRESSABLE_FLAG == 'PO_CANCELLED') & (df.PACKOUT_QUANTITY.notnull()) & (
        df.ORGANIZATION_CODE.isin(org_list))].copy()
    df_not_cancelled = df[(df.ADDRESSABLE_FLAG != 'PO_CANCELLED') & (df.ASN_CREATION_DATE.isnull()) & (
        df.ORGANIZATION_CODE.isin(org_list))].copy()

    # Pivot them
    df_rtv_p = df_rtv.pivot_table(index=['PO_NUMBER'], columns='PRODUCT_ID',
                                  values='ORDERED_QUANTITY', aggfunc=sum)
    df_not_cancelled_p = df_not_cancelled.pivot_table(index=['PO_NUMBER'], columns='PRODUCT_ID',
                                                      values='ORDERED_QUANTITY', aggfunc=sum)
    # scale down each PO to 1 unit
    df_rtv_p = df_rtv_p.apply(lambda x: x / x.min(), axis=1)
    df_not_cancelled_p = df_not_cancelled_p.apply(lambda x: x / x.min(), axis=1)

    # Create dic for both
    rtv_dic = create_dic(df_rtv_p)
    not_cancelled_dic = create_dic(df_not_cancelled_p)

    '''
    # I) Compare these two dic & create new col to indicate the matching cancelled orders for good backlog orders
    val_dic = compare_input_and_ref_dic(not_cancelled_dic, rtv_dic)
    df.loc[:, 'rtv_po_with_same_config'] = df.PO_NUMBER.map(
        lambda x: val_dic.get(x) if val_dic.get(x) != [] else np.nan)
    '''

    # II) Compare these two dic the other way & create new col to indicate the matching good backlog orders for rtv orders
    val_dic = compare_input_and_ref_dic(rtv_dic, not_cancelled_dic)
    df.loc[:, 'good_po_with_same_config'] = df.PO_NUMBER.map(
        lambda x: val_dic.get(x) if val_dic.get(x) != [] else np.nan)

    del val_dic,not_cancelled_dic,rtv_dic,df_not_cancelled_p,df_rtv_p,df_not_cancelled,df_rtv
    gc.collect()


    return df



def add_rtv_good_po_qty_col(df):
    '''
    Add a col to indicate the total qty from the same config PO
    :param df:
    :return:
    '''

    dfx=df[df.need_rtv=='YES'].copy()
    rtv_dic={}

    regex=re.compile(r'\d{8,10}-\d{1,2}')

    for rtv_po,good_po in zip(dfx.PO_NUMBER,dfx.good_po_with_same_config):
        if regex.search(str(good_po))!=None:
            good_po_list=regex.findall(str(good_po))
        else:
            good_po_list=[]

        rtv_dic[rtv_po]=good_po_list

    for rtv_po,good_po_list in rtv_dic.items():
        good_po_qty=df[(df.PO_NUMBER.isin(good_po_list))].ORDERED_QUANTITY.sum()
        rtv_dic[rtv_po]=good_po_qty

    df.loc[:,'good_po_qty']=df.PO_NUMBER.map(lambda x: rtv_dic.get(x))

    del rtv_dic
    gc.collect()

    return df




def exception_check_and_recommendation(df):
    '''
    Make exception check of data and make recommendation of actions; create one col to store the info.
    :param df:
    :return: df
    '''

    # Calculate MCD first so else using np.nan (any better way to avoid using np.nan?)
    df.loc[:, 'exception_highlight'] = df[(df.MCD.notnull()) & (df.MFG_HOLD == 'N')]['MCD'].map(
        lambda x: 'MCD w/o waiver' if not 'WAIVER' in str(x).upper() else '')

    # Brazil operating unit orders
    df.loc[:, 'exception_highlight'] = np.where(df.SALES_ORDER_OPERATING_UNIT == 'CISCO BRAZIL CA OPERATING UNIT',
                                                     np.where(df.exception_highlight.isnull(),
                                                              'TSP order, convert or cancel',
                                                              'TSP order, convert or cancel' + '/' + df.exception_highlight),
                                                     df.exception_highlight)

    # Missing shipping route code
    df.loc[:, 'exception_highlight'] = np.where(df.OTM_SHIPPING_ROUTE_CODE.isnull(),
                                                     np.where(df.exception_highlight.isnull(),
                                                              'Missing shipping route code',
                                                              'Missing shipping route code' + '/' + df.exception_highlight),
                                                     df.exception_highlight)


    return df

def read_3a4(fname_3a4):
    '''
    Read full 3a4 dataset without parsing dates
    '''

    ext = os.path.splitext(fname_3a4)[1]

    # formally read 3a4
    if ext == '.csv':
        df_3a4 = pd.read_csv(fname_3a4, low_memory=False,encoding='ISO-8859-1')
    else:
        df_3a4=pd.read_excel(fname_3a4,)

    file_size=round(os.path.getsize(fname_3a4)/1024000,1)
    df_rows=df_3a4.shape[0]

    print('3A4 file read (size: {}Mb; rows: {}'.format(str(file_size),str(df_rows)))

    return df_3a4


def read_3a4_parse_dates(fname_3a4, date_col):
    '''
    Check format and read 3a4
    :param fname_3a4:
    :param date_col:
    :param ext: extension of the filename
    :param original_fname: original uploaded filename without processing
    :return:
    '''

    ext = os.path.splitext(fname_3a4)[1]

    # formally read 3a4
    if ext == '.csv':
        df_3a4 = pd.read_csv(fname_3a4, low_memory=False,
                         # nrows=5,
                         parse_dates=date_col,
                         encoding='ISO-8859-1',
                         )
    else:
        df_3a4=pd.read_excel(fname_3a4,
                             parse_dates=date_col)

    file_size=round(os.path.getsize(fname_3a4)/1024000,1)
    df_rows=df_3a4.shape[0]

    print('3A4 file read (size: {}Mb; rows: {}'.format(str(file_size),str(df_rows)))

    return df_3a4

def read_3a4_header_to_check_format(fname_3a4,nrows=3):
    '''
    Check format and read 3a4
    '''

    ext = os.path.splitext(fname_3a4)[1]

    # formally read 3a4
    if ext == '.csv':
        df_3a4 = pd.read_csv(fname_3a4, low_memory=False,
                         nrows=nrows,
                         encoding='ISO-8859-1',
                         )
    else:
        df_3a4=pd.read_excel(fname_3a4,nrows=nrows)

    file_size=round(os.path.getsize(fname_3a4)/1024000,1)
    df_rows=df_3a4.shape[0]

    print('3A4 file read (size: {}Mb; rows: {}'.format(str(file_size),str(df_rows)))

    return df_3a4


def basic_data_processing_global(df_3a4,region,org_name_global):
    # exclude lines with BU or PF as null
    df_3a4 = df_3a4[(df_3a4.BUSINESS_UNIT.notnull()) & (df_3a4.PRODUCT_FAMILY.notnull())].copy()
    df_3a4 = df_3a4[df_3a4.ORGANIZATION_CODE.isin(org_name_global[region])].copy()

    # 生成main_X列 - discarded due to now based on 3a4 with main PID only
    df_3a4 = commonize_and_create_main_item(df_3a4, 'BUSINESS_UNIT', 'main_bu')
    df_3a4 = commonize_and_create_main_item(df_3a4, 'PRODUCT_FAMILY', 'main_pf')

    # update CRBU name: main_bu changed to "CRBU-NG" or "CRBU-other"
    df_3a4 = update_crbu_bu_name(df_3a4)

    # create po_rev column - this is must since po_rev_unstg is used in other sub programs
    df_3a4 = create_po_rev_unstg_col(df_3a4)

    # Redefine the addressable flag, add in MFG_HOLD, and split out wk+1, wk+2 for outside_window portion
    df_3a4 = redefine_addressable_flag_new(df_3a4, mfg_holds)

    return df_3a4


def basic_data_processin_dfpm_app(df_3a4):

    # exclude lines with BU or PF as null
    df_3a4 = df_3a4[(df_3a4.BUSINESS_UNIT.notnull()) & (df_3a4.PRODUCT_FAMILY.notnull())].copy()

    # Redefine the addressable flag, add in MFG_HOLD, and split out wk+1, wk+2 for outside_window portion
    df_3a4 = redefine_addressable_flag_new(df_3a4, mfg_holds)

    # 生成main_X列
    df_3a4 = commonize_and_create_main_item(df_3a4, 'BUSINESS_UNIT', 'main_bu')
    df_3a4 = commonize_and_create_main_item(df_3a4, 'PRODUCT_FAMILY', 'main_pf')

    # update CRBU main_bu name: PF 8000& SPFIRE to CRBU-NG and rest to CRBU-other
    df_3a4=update_crbu_bu_name(df_3a4)

    # reformat 3a4 comments for easier reading
    df_3a4 = reformat_and_keep_latest_3a4_comment(df_3a4)

    # 提取category from comments
    df_3a4 = extract_category_from_3a4(df_3a4)

    # Create exception_highlight col
    df_3a4 = exception_check_and_recommendation(df_3a4)

    # create po_rev column
    df_3a4=create_po_rev_unstg_col(df_3a4)

    return df_3a4


def basic_data_processing_backlog_ranking(df_3a4,org):
    """
    Basica data processing for the ranking report - used by CM
    """

    df_3a4 = df_3a4[df_3a4.ORGANIZATION_CODE==org].copy()

    if df_3a4.shape[0]>0:
        # Redefine the addressable flag, add in MFG_HOLD,
        df_3a4 = redefine_addressable_flag_new(df_3a4, mfg_holds)

        # reformat 3a4 comments for easier reading
        df_3a4 = reformat_and_keep_latest_3a4_comment(df_3a4)

    return df_3a4

def update_crbu_bu_name(df_3a4):
    """
    Update CRBU BU name to separate out PF 8000 and SPFIRE to CRBU-NG; keep NCS5500 and others as CRBU unchanged.
    :param df_3a4:
    :return:
    """
    df_3a4.loc[:,'main_bu']=np.where(df_3a4.main_bu=='CRBU',
                                            np.where(df_3a4.main_pf.isin(['8000','SPFIRE']),
                                                   'CRBU-NG',
                                                   'CRBU-other'),
                                           df_3a4.main_bu)

    return df_3a4

def create_rtv_config_col(df_3a4):
    '''
    Create new col to indicate same config orders
    :return: df_3a4
    '''
    # only do for selected org to save running time
    df_3a4 = create_config_matching_rtv_col(df_3a4, ['FOC', 'FDO'])
    df_3a4=add_rtv_good_po_qty_col(df_3a4)

    return df_3a4

def send_top_customer_booking_by_email(region, data,threshold,login_user,to_address,sender):
    """
    Send the result by email
    """
    subject = region + ' top customers and bookings summary (sent by: '+login_user +')'
    html = 'top_customer_and_booking_email.html'

    send_attachment_and_embded_image(to_address, subject, html,
                                     sender=sender,
                                     att_filenames=None,
                                     bcc=[super_user + '@cisco.com'],
                                     data=data,
                                     threshold=threshold)

def create_booking_and_backlog_customer_summary(df_3a4_main,region):
    """
    Create summary for the top$ backlog customers and recent bookings for them
    """
    # bookings by site
    site_booking_summary=[]
    dfp=df_3a4_main.pivot_table(index=['ORGANIZATION_CODE'], columns='LINE_CREATION_DATE',
                                  values='po_rev_unstg', aggfunc=sum) / 1000000
    dfp.columns = dfp.columns.map(lambda x: x.strftime('%m-%d'))
    dfp.loc[:, 'Total booking'] = dfp.iloc[:,-top_customers_bookings_history_days:].sum(axis=1)
    dfp.loc[:, 'Total backlog'] = dfp.sum(axis=1) - dfp['Total booking']
    dfp = dfp.iloc[:, -(top_customers_bookings_history_days + 2):].copy()
    dfp.sort_values(by='Total booking',ascending=False,inplace=True)
    dfp.loc['Total', :] = dfp.sum(axis=0)
    print(dfp)


    dfp.fillna(0, inplace=True)
    dfp = dfp.applymap(lambda x: round(x))
    dfp.reset_index(inplace=True)
    dfp.rename(columns={'ORGANIZATION_CODE': 'Org code'}, inplace=True)
    site_booking_summary.append((dfp.columns, dfp.values))


    # collect top backlog customers by org - sort by total backlog and PO detail sort by creationg date
    top_backlog_customer_summary = []  #
    df_3a4_main.loc[:, 'END_CUSTOMER_NAME'] = np.where(df_3a4_main.END_CUSTOMER_NAME.isnull(),
                                                       'dummy',
                                                       df_3a4_main.END_CUSTOMER_NAME)
    dfp = df_3a4_main.pivot_table(index=['ORGANIZATION_CODE', 'END_CUSTOMER_NAME'], columns='LINE_CREATION_DATE',
                                  values='po_rev_unstg', aggfunc=sum) / 1000000
    dfp.columns = dfp.columns.map(lambda x: x.strftime('%m-%d'))
    dfp.loc[:, 'Total backlog'] = dfp.sum(axis=1)
    dfp = dfp.iloc[:, -(top_customers_bookings_history_days+1):].copy()

    for org in org_name_global[region][region]:
        dfp_org = dfp.loc[(org, slice(None)), :].copy()
        dfp_org.sort_values(by='Total backlog', ascending=False, inplace=True)
        dfp_org.loc[(org, 'Total'), :] = dfp_org.sum(axis=0)
        dfp_org.reset_index(inplace=True)
        dfp_org=dfp_org[(dfp_org.END_CUSTOMER_NAME!='dummy')&(dfp_org['Total backlog'] >= top_customers_bookings_threshold)].copy()
        dfp_org.set_index(['ORGANIZATION_CODE', 'END_CUSTOMER_NAME'],inplace=True)
        dfp_org = dfp_org.applymap(lambda x: round(x, 1))
        dfp_org.fillna('', inplace=True)

        # find the top PO and put into the last PO detail col
        if dfp_org.shape[0] > 1:  # more than the total record
            customer_list = [x[1] for x in dfp_org.index]
            for customer in customer_list:
                dfp_org_cus = df_3a4_main[
                    (df_3a4_main.ORGANIZATION_CODE == org) & (df_3a4_main.END_CUSTOMER_NAME == customer)].copy()
                #dfp_org_cus.sort_values(by='po_rev_unstg', ascending=False, inplace=True)
                dfp_org_cus.sort_values(by='LINE_CREATION_DATE', ascending=False, inplace=True)

                top_po_list = []
                for row in dfp_org_cus.itertuples():
                    bu=row.BUSINESS_UNIT
                    rev=str(round(row.po_rev_unstg / 1000000, 1))
                    try:
                        create_date=pd.to_datetime(row.LINE_CREATION_DATE).strftime('%m-%d')
                    except:
                        create_date='N/A'
                    try:
                        fcd=pd.to_datetime(row.CURRENT_FCD_NBD_DATE).strftime('%m-%d')
                    except:
                        fcd='N/A'

                    top_po_list.append(
                        row.PO_NUMBER + '(' + bu + ', ' + rev + 'm, Enter date:' + create_date + ', FCD:' + fcd + ')')

                dfp_org.loc[(org, customer), 'PO Details'] = '  '.join(top_po_list)

            dfp_org.reset_index(inplace=True)
            dfp_org.rename(columns={'ORGANIZATION_CODE': 'Org Code'}, inplace=True)
            top_backlog_customer_summary.append((dfp_org.columns, dfp_org.values))

    return site_booking_summary,top_backlog_customer_summary

def create_top_customer_and_booking_summary(df_3a4_main,region):
    """
    Create summary for the top$ backlog customers and recent bookings for them
    """

    # top bookings by customers
    dfp = df_3a4_main.pivot_table(index=['ORGANIZATION_CODE', 'END_CUSTOMER_NAME'], columns='LINE_CREATION_DATE',
                                  values='po_rev_unstg', aggfunc=sum) / 1000000
    dfp.columns = dfp.columns.map(lambda x: x.strftime('%m-%d'))
    dfp.loc[:, 'Total backlog'] = dfp.sum(axis=1)
    dfp = dfp.iloc[:, -(top_customers_bookings_history_days+1):].copy()

    # collect the top bookings customers
    dfp.loc[:, 'Total booking']=dfp.sum(axis=1)
    dfp.loc[:, 'Total booking']=dfp['Total booking']-dfp['Total backlog'] # above sum also incldued the totla backlog, thus deduct it\

    top_booking_customer_summary = [] # sort by total booking in the past x days

    dfp.sort_values(by='Total booking', ascending=False, inplace=True)
    dfp_booking = dfp[dfp['Total booking'] >= top_customers_bookings_threshold].copy()
    #dfp_booking.drop('Total booking', axis=1, inplace=True)
    dfp_booking = dfp_booking.applymap(lambda x: round(x, 1))
    dfp_booking.fillna('', inplace=True)
    dfp_booking.reset_index(inplace=True)
    dfp_booking.rename(columns={'ORGANIZATION_CODE': 'Org code'}, inplace=True)
    top_booking_customer_summary.append((dfp_booking.columns, dfp_booking.values))


    # collect top backlog customers by org
    dfp.drop('Total booking', axis=1, inplace=True)
    top_backlog_customer_summary = []  # sort by total backlog

    for org in org_name_global[region][region]:
        dfp_org = dfp.loc[(org, slice(None)), :].copy()
        dfp_org.sort_values(by='Total backlog', ascending=False, inplace=True)
        dfp_org.loc[(org, 'Total'), :] = dfp_org.sum(axis=0)
        dfp_org = dfp_org[dfp_org['Total backlog'] >= top_customers_bookings_threshold]
        dfp_org = dfp_org.applymap(lambda x: round(x, 1))
        dfp_org.fillna('', inplace=True)

        # find the top PO and put into the last PO detail col
        if dfp_org.shape[0] > 1:  # more than the total record
            customer_list = [x[1] for x in dfp_org.index]
            for customer in customer_list:
                dfp_org_cus = df_3a4_main[
                    (df_3a4_main.ORGANIZATION_CODE == org) & (df_3a4_main.END_CUSTOMER_NAME == customer)].copy()
                #dfp_org_cus.sort_values(by='po_rev_unstg', ascending=False, inplace=True)
                dfp_org_cus.sort_values(by='LINE_CREATION_DATE', ascending=False, inplace=True)

                top_po_list = []
                for row in dfp_org_cus.itertuples():
                    bu=row.BUSINESS_UNIT
                    rev=str(round(row.po_rev_unstg / 1000000, 1))
                    try:
                        create_date=pd.to_datetime(row.LINE_CREATION_DATE).strftime('%m-%d')
                    except:
                        create_date='N/A'
                    try:
                        fcd=pd.to_datetime(row.CURRENT_FCD_NBD_DATE).strftime('%m-%d')
                    except:
                        fcd='N/A'

                    top_po_list.append(
                        row.PO_NUMBER + '(' + bu + ', ' + rev + 'm, Enter date:' + create_date + ', FCD:' + fcd + ')')

                dfp_org.loc[(org, customer), 'PO Details'] = '  '.join(top_po_list)

            dfp_org.reset_index(inplace=True)
            dfp_org.rename(columns={'ORGANIZATION_CODE': 'Org Code'}, inplace=True)
            top_backlog_customer_summary.append((dfp_org.columns, dfp_org.values))

    return top_booking_customer_summary,top_backlog_customer_summary

def create_and_send_wnbu_compliance(wnbu_compliance_hold_emails, df_compliance_release, df_compliance_hold,
                                    df_country_missing, df_compliance_table, login_user,sender):
    to_address = wnbu_compliance_hold_emails
    to_address.append(login_user + '@cisco.com')
    subject = 'WNBU Compliance hold status summary (sent by: '+login_user +')'
    html_template = 'wnbu_compliance.html'

    msg,size_over_limit = send_attachment_and_embded_image(to_address, subject, html_template,
                                                           sender=sender,
                                                           bcc=None,
                                                           ok_release_header=df_compliance_release.columns,
                                                           ok_release_data=df_compliance_release.values,
                                                           keep_on_hold_header=df_compliance_hold.columns,
                                                           keep_on_hold_data=df_compliance_hold.values,
                                                           country_missing_header=df_country_missing.columns,
                                                           country_missing_data=df_country_missing.values,
                                                           compliance_table_header=df_compliance_table.columns,
                                                           compliance_table_data=df_compliance_table.values,
                                                           att_filenames=None,
                                                           embeded_filenames=None)


def add_cm_ctb_to_3a4(df_3a4):
    """
    Read CM CTB from smartsheet and add to 3a4
    """
    ctb_df, ctb_error_msg = read_ctb_from_smartsheet()

    regex_line = re.compile(r'-\d+')

    df_3a4.loc[:, 'line'] = df_3a4.PO_NUMBER.map(lambda x: regex_line.search(x).group())
    df_3a4.loc[:, 'SO_SS_LN'] = df_3a4.SO_SS + df_3a4.line

    df_3a4 = pd.merge(df_3a4, ctb_df, left_on='SO_SS_LN', right_on='SO_SS_LN', how='left')

    df_3a4.drop(['line', 'SO_SS_LN'], axis=1, inplace=True)

    del ctb_df
    gc.collect()

    return df_3a4,ctb_error_msg

def create_ctb_summaries(df_3a4):
    # Create addressable+CTB summary (for adding into the spreadsheets for dfpm)

    addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf = create_addressable_summary_with_ctb(df_3a4)
    ctb_summary_for_material = create_ctb_summary_for_material_gating(df_3a4)


    return addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf, ctb_summary_for_material


def create_addressable_summary_and_comb_addressable_history(df_3a4, org_name_region, region, addr_fname):
    '''
    Create addressable summary for later use (add to 3a4 or making charts); collect addressable to tracker, and create
    the tracker data for later use (add to email, or making trending chart)
    '''
    #print('saving file')
    #df_3a4.to_csv('3a4 processed.csv',index=False)

    #print('File saved')

    # create addressable summary - list of df
    addr_df_summary = create_addr_summary(df_3a4, org_name_region)

    # Add new addressable data to existing tracker data
    if region=='APJC':
        addr_df_dict,df_apjc,df_foc,df_fdo,df_shk,df_jpe,df_ncb = collect_new_addr_and_comb_historical_apjc(addr_df_summary, addr_fname)
    elif region=='EMEA':
        addr_df_dict,df_emea,df_fcz,df_fve = collect_new_addr_and_comb_historical_emea(addr_df_summary, addr_fname)
    elif region=='Americas':
        addr_df_dict,df_americas,df_ftx,df_tau,df_sjz,df_fgu,df_jmx,df_fjz,df_tsp = collect_new_addr_and_comb_historical_americas(addr_df_summary, addr_fname)


    return addr_df_summary, addr_df_dict

def create_and_save_regional_3a4(region,df_3a4,addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf, ctb_summary_for_material,df_asp):
    """
    Create and save a APJC 3a4
    """

    col=[]
    for x in col_3a4_regional:
        if x in df_3a4.columns:
            col.append(x)

    df_dict={}
    df_3a4=df_3a4[col]
    df_3a4.set_index('ORGANIZATION_CODE',inplace=True)
    df_dict['3a4'] = df_3a4
    df_dict['addr_ctb_bu'] = addr_ctb_by_org_bu
    df_dict['addr_ctb_pf'] = addr_ctb_by_org_bu_pf
    df_dict['ctb_material'] = ctb_summary_for_material
    df_dict['asp'] = df_asp

    file_path = base_dir_output
    fname = region + ' 3a4 ' + pd.Timestamp.now().strftime('%m-%d %H:%M') + '.xlsx'
    full_path = os.path.join(file_path, fname)
    write_excel_file(full_path, df_dict)



def save_addr_tracker(df_3a4,addr_df_dict, region, org_name,addr_fname):
    '''
    Save new tracker - Only when add_to_tracker and also 3a4 include all APJC org then write the addressable into tracker
    :param org_3a4:
    :param addr_df_dict:
    :param add_to_tracker:
    :return:
    '''

    org_3a4=df_3a4.ORGANIZATION_CODE.unique()
    if region=='APJC':
        if np.all(np.in1d(org_name['APJC'],org_3a4)):
            write_excel_file(addr_fname, addr_df_dict)
            msg = 'New addressable data added to tracker.'
            print(msg)
        else:
            msg = 'Addresable backlog tracker not added - not all APJC orgs are included in the file.'
            print(msg)
    elif region=='EMEA':
        if np.all(np.in1d(org_name['EMEA'],org_3a4)):
            write_excel_file(addr_fname, addr_df_dict)
            msg = 'New addressable data added to tracker.'
            print(msg)
        else:
            msg = 'Addresable backlog tracker not added - not all APJC orgs are included in the file.'
            print(msg)
    elif region=='Americas':
        if np.all(np.in1d(org_name['Americas'],org_3a4)):
            write_excel_file(addr_fname, addr_df_dict)
            msg = 'New addressable data added to tracker.'
            print(msg)
        else:
            msg = 'Addresable backlog tracker not added - not all APJC orgs are included in the file.'
            print(msg)



def create_and_send_addressable_summaries(top_booking_customer_summary,top_customers_bookings_history_days,top_customers_bookings_threshold,addr_df_summary,
                                          addr_df_dict, org_name_region, backlog_dashboard_emails,region,sender, login_user):
    '''
    Create the addressable tracker,snapshot chart, trending chart, and send by email.
    '''

    # Below create charts
    # Create addressable chart - for APJC by Org
    if region=='APJC':
        addr_chart_file_name=backlog_chart_global[region]['apjc_add_summary']
    elif region=='EMEA':
        addr_chart_file_name = backlog_chart_global[region]['emea_add_summary']
    elif region == 'Americas':
        addr_chart_file_name = backlog_chart_global[region]['americas_add_summary']


    # create regional chart: with org=None
    create_addr_charts(addr_df_summary, org_name_region, addr_chart_file_name, region, org=None)

    # Create addressable chart - for each org by BU
    for org in org_name_global[region]:
        if org != 'APJC' and org != 'EMEA' and org != 'Americas':
            addr_chart_file_name = backlog_chart_global[region][org.lower() + '_add_summary']
            create_addr_charts(addr_df_summary, org_name_region, addr_chart_file_name, region, org=org)

    # create_addr_trending_chart
    create_addr_trending_chart(addr_df_dict,backlog_chart_global[region])

    # Send addressable backlog charts to defined users
    to_address = backlog_dashboard_emails
    to_address.append(login_user + '@cisco.com')
    subject = region + ' backlog summaries (sent by: ' + login_user + ')'

    if region=='APJC':
        html = 'apjc_backlog_summary.html'

        msg,size_over_limit = send_attachment_and_embded_image(to_address, subject, html, att_filenames=None,
                                                                sender=sender,
                                                                bcc=[super_user + '@cisco.com'],
                                                                embeded_filenames=backlog_chart_global[region],
                                                                banner_addr='cid:banner_addr',
                                                                apjc_add_summary='cid:apjc_add_summary',
                                                                foc_add_summary='cid:foc_add_summary',
                                                                fdo_add_summary='cid:fdo_add_summary',
                                                                jpe_add_summary='cid:jpe_add_summary',
                                                                shk_add_summary='cid:shk_add_summary',
                                                                ncb_add_summary='cid:ncb_add_summary',
                                                                apjc_add_trending='cid:apjc_add_trending',
                                                                foc_add_trending='cid:foc_add_trending',
                                                                fdo_add_trending='cid:fdo_add_trending',
                                                                jpe_add_trending='cid:jpe_add_trending',
                                                                shk_add_trending='cid:shk_add_trending',
                                                                ncb_add_trending='cid:ncb_add_trending',
                                                               top_booking_customer_summary=top_booking_customer_summary,
                                                               top_customers_bookings_history_days=top_customers_bookings_history_days,
                                                               threshold=top_customers_bookings_threshold,
                                                            )

    elif region=='EMEA':
        html = 'emea_backlog_summary.html'

        msg, size_over_limit = send_attachment_and_embded_image(to_address, subject, html, att_filenames=None,
                                                                sender=sender,
                                                                bcc=[super_user + '@cisco.com'],
                                                                embeded_filenames=backlog_chart_global[region],
                                                                banner_addr='cid:banner_addr',
                                                                emea_add_summary='cid:emea_add_summary',
                                                                fcz_add_summary='cid:fcz_add_summary',
                                                                fve_add_summary='cid:fve_add_summary',
                                                                emea_add_trending='cid:emea_add_trending',
                                                                fcz_add_trending='cid:fcz_add_trending',
                                                                fve_add_trending='cid:fve_add_trending',
                                                                top_booking_customer_summary=top_booking_customer_summary,
                                                                top_customers_bookings_history_days=top_customers_bookings_history_days,
                                                                threshold = top_customers_bookings_threshold,
                                                                )
    elif region=='Americas':
        html = 'americas_backlog_summary.html'

        msg, size_over_limit = send_attachment_and_embded_image(to_address, subject, html, att_filenames=None,
                                                                sender=sender,
                                                                bcc=[super_user + '@cisco.com'],
                                                                embeded_filenames=backlog_chart_global[region],
                                                                banner_addr='cid:banner_addr',
                                                                americas_add_summary='cid:americas_add_summary',
                                                                ftx_add_summary='cid:ftx_add_summary',
                                                                tau_add_summary='cid:tau_add_summary',
                                                                sjz_add_summary='cid:sjz_add_summary',
                                                                fgu_add_summary='cid:fgu_add_summary',
                                                                jmx_add_summary='cid:jmx_add_summary',
                                                                fjz_add_summary='cid:fjz_add_summary',
                                                                tsp_add_summary='cid:tsp_add_summary',
                                                                americas_add_trending='cid:americas_add_trending',
                                                                ftx_add_trending='cid:ftx_add_trending',
                                                                tau_add_trending='cid:tau_add_trending',
                                                                sjz_add_trending='cid:sjz_add_trending',
                                                                fgu_add_trending='cid:fgu_add_trending',
                                                                jmx_add_trending='cid:jmx_add_trending',
                                                                fjz_add_trending='cid:fjz_add_trending',
                                                                tsp_add_trending='cid:tsp_add_trending',
                                                                top_booking_customer_summary=top_booking_customer_summary,
                                                                top_customers_bookings_history_days=top_customers_bookings_history_days,
                                                                threshold=top_customers_bookings_threshold,
                                                                )



def create_and_send_apjc_outlier_summaries(df_3a4, outlier_elements, outlier_chart_apjc, apjc_outlier_emails,
                                      org_name,comparison_days, login_user):
    '''
    Create the outlier summaries and send via email
    :param df_3a4:
    :param outlier_elements:
    :param outlier_chart_apjc:
    :param apjc_outlier_emails:
    :return:
    '''
    # Create the outlier chart and spreadsheet - for APJC aggregaged
    create_outlier_chart(df_3a4, 'APJC', outlier_elements, outlier_chart_apjc.values())
    # create the outlier df
    df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel, df_partial_staged,df_missed_recommit \
        = create_outlier_df(df_3a4, outlier_elements)

    # create outlier change comparison df
    df_comparison=calculate_outlier_changes(outlier_elements, 'APJC', df_3a4, org_name,comparison_days)

    # create the APJC outlier spreadsheet
    fname = 'APJC outlier summary.xlsx'
    full_path = os.path.join(base_dir_output,fname)
    create_outlier_spreadsheet(full_path, df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel,
                               df_partial_staged,df_missed_recommit)

    att_files_outlier = [(base_dir_output, fname)]  # List of tuples (path, file_name)

    # reset index to have ORGANIZATION_CODE col out of index
    df_not_booked.reset_index(inplace=True)
    df_not_scheduled.reset_index(inplace=True)
    df_not_packed.reset_index(inplace=True)
    df_partial_staged.reset_index(inplace=True)
    df_missed_recommit.reset_index(inplace=True)
    df_aging_cancel.reset_index(inplace=True)

    # Send outlier charts and spreadsheet to defined users
    to_address = apjc_outlier_emails
    subject = 'APJC Outlier summaries (sent by: '+login_user +')'
    html = 'outlier.html'
    top_x = outlier_detail_top_x

    msg,size_over_limit = send_attachment_and_embded_image(to_address, subject, html,
                                           att_filenames=att_files_outlier,
                                           embeded_filenames=outlier_chart_apjc,
                                           outlier_book='cid:outlier_book_apjc',
                                           outlier_schedule='cid:outlier_schedule_apjc',
                                           outlier_pack='cid:outlier_pack_apjc',
                                           outlier_cancel='cid:outlier_cancel_apjc',
                                           outlier_partial='cid:outlier_partial_apjc',
                                           top_x=top_x,
                                           not_booked=df_not_booked.values[:top_x],
                                           not_scheduled=df_not_scheduled.values[:top_x],
                                           not_packed=df_not_packed.values[:top_x],
                                           aging_cancel=df_aging_cancel.values[:top_x],
                                           partial_staged=df_partial_staged.values[:top_x],
                                           #missed_recommit=df_missed_recommit.values[:top_x],
                                           comparison_days=comparison_days,
                                           outlier_comparison=df_comparison.values,
                                               china_outlier=False
                                           )

    return df_not_booked, df_not_scheduled, df_not_packed,df_aging_cancel,df_partial_staged, df_missed_recommit

def create_and_send_cm_3a4(df_3a4, cm_emails_to, outlier_elements,outlier_chart_foc,outlier_chart_fdo,
                                                     org_name,outlier_comparison, login_user):
    '''
    Create 3a4 by CM and send via email
    :param df_3a4:
    :param email_to_only: send email to this if this is not blank
    :param cm_emails:
    :param outlier_elements: for creating the outlier df
    :param comparison_days: days back to compare outlier
    :return: None
    '''

    # create the outlier df - combined all orgs
    df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel, df_partial_staged,df_missed_recommit \
        = create_outlier_df(df_3a4, outlier_elements)


    # create outlier charts by org: now only cover FOC/FDO
    for org in cm_emails_to.keys():
        if org == 'FOC':
            outlier_chart_fname_list = outlier_chart_foc.values()
        elif org == 'FDO':
            outlier_chart_fname_list = outlier_chart_fdo.values()

        # Create the outlier chart
        create_outlier_chart(df_3a4, org, outlier_elements, outlier_chart_fname_list)

    # create outlier change comparison df, spreadsheet that include 3a4 and outlier by org, and send by email by org
    att_files_cm = {}
    for org in cm_emails_to.keys():
        # create outlier change comparison df
        df_comparison = calculate_outlier_changes(outlier_elements, org, df_3a4, org_name,outlier_comparison)

        # create spreadsheets
        dfx = df_3a4[(df_3a4.ORGANIZATION_CODE == org)].copy()
        #col = np.intersect1d(col_3a4_cm, df_3a4.columns)
        dfx=dfx[col_3a4_cm].copy()
        dfx.set_index('ORGANIZATION_CODE',inplace=True)

        file_path = base_dir_output
        fname = org + '_3a4_and_outlier_file ' + pd.Timestamp.now().strftime('%m-%d %H:%M') + '.xlsx'
        full_path = os.path.join(file_path, fname)

        # select df base on org
        #df_not_booked_x = df_not_booked[df_not_booked.ORGANIZATION_CODE == org]
        df_not_scheduled_x = df_not_scheduled[df_not_scheduled.ORGANIZATION_CODE == org]
        df_not_packed_x = df_not_packed[df_not_packed.ORGANIZATION_CODE == org]
        df_partial_staged_x = df_partial_staged[df_partial_staged.ORGANIZATION_CODE == org]
        df_aging_cancel_x = df_aging_cancel[df_aging_cancel.ORGANIZATION_CODE == org]
        df_missed_recommit_x=df_missed_recommit[df_missed_recommit.ORGANIZATION_CODE==org]

        df_dict = {'processed_3a4': dfx}

        # Write multiple DF to excel based on df_dict
        write_excel_file(full_path, df_dict)

        # 添加文件到附件列表
        att_files_cm[org] = [(file_path, fname)]  # List of tuples (path, file_name)

        # Send org level 3a4 with ranking to to CM
        to_address = cm_emails_to[org]
        to_address.append(login_user + '@cisco.com')
        subject = org + ' 3a4 and outlier backlog summaries  (sent by: '+login_user +')'
        html = 'outlier.html'
        top_x = outlier_detail_top_x

        if org=='FOC':
            embeded_charts=outlier_chart_foc
        elif org=='FDO':
            embeded_charts = outlier_chart_fdo

        msg,size_over_limit = send_attachment_and_embded_image(to_address, subject, html,
                                               att_filenames=att_files_cm[org],
                                               embeded_filenames=embeded_charts,
                                               outlier_schedule='cid:outlier_schedule' + '_' + org.lower(),
                                               outlier_pack='cid:outlier_pack' + '_' + org.lower(),
                                               outlier_cancel='cid:outlier_cancel' + '_' + org.lower(),
                                               outlier_partial='cid:outlier_partial' + '_' + org.lower(),
                                               outlier_recommit='cid:outlier_recommit'+ '_' + org.lower(),
                                               top_x=top_x,
                                               not_scheduled=df_not_scheduled_x.values[:top_x],
                                               not_packed=df_not_packed_x.values[:top_x],
                                               aging_cancel=df_aging_cancel_x.values[:top_x],
                                               partial_staged=df_partial_staged_x.values[:top_x],
                                               missed_recommit=df_missed_recommit_x.values, # list all, instead of top_x only
                                               comparison_days=outlier_comparison,
                                               outlier_comparison=df_comparison.values,
                                                   china_outlier=True
                                               )

    del df_comparison,df_not_scheduled_x,df_not_packed_x,df_aging_cancel_x,df_partial_staged_x,df_missed_recommit_x
    gc.collect()




def create_and_send_3a4_backlog_ranking(df_3a4, to_address, org, login_user,login_name):
    '''
    Create 3a4 backlog ranking file and send via email
    '''

    # create the output file
    df_3a4 = df_3a4[col_3a4_backlog_ranking_output_col].copy()
    df_3a4.set_index('ORGANIZATION_CODE', inplace=True)
    fname=org + ' 3a4 backlog ranking ' + login_user + ' ' + pd.Timestamp.now().strftime('%m-%d %H:%M') + '.xlsx'
    output_file = os.path.join(base_dir_output, fname)
    df_3a4.to_excel(output_file)

    # 添加文件到附件列表
    att_files=[(base_dir_output, fname)]  # List of tuples (path, file_name)
    subject = org + ' 3a4 backlog ranking'
    html = 'backlog_ranking_email.html'

    msg,size_over_limit = send_attachment_and_embded_image(to_address, subject, html,
                                                           sender=login_name + ' via DFPM automation tool',
                                                            att_filenames=att_files)


def generate_dfpm_mapping_dict(df_dfpm_mapping):
    """
    Generate the mapping dict based on mapping database df:df_dfpm_mapping
    """
    dfpm_mapping = {}

    for row in df_dfpm_mapping.itertuples():
        dfpm = row.DFPM
        org = row.Org
        bu_list = row.BU.split('/')
        extra_pf_list = row.Extra_PF.split('/')
        exclusion_pf_list = row.Exclusion_PF.split('/')
        if dfpm not in dfpm_mapping.keys():
            coverage = {}
        coverage[org] = (bu_list, extra_pf_list, exclusion_pf_list)
        dfpm_mapping[dfpm] = coverage

    return dfpm_mapping


def reformat_and_keep_latest_3a4_comment(df_3a4):
    '''
    To remove the old 3a4 comments and only keep the latest one (on the top) - based on "\n***" processed comments
    :param df_3a4:
    :return:
    '''

    df_3a4.loc[:, 'COMMENTS'] = df_3a4.COMMENTS.map(lambda x: str(x).replace('[', '\n***[') if (x != np.nan and x!='') else '')

    df_3a4.loc[:, 'COMMENTS'] = df_3a4.COMMENTS.map(lambda x: str(x).split('***')[1] if '***' in x else x)

    df_3a4.loc[:, 'COMMENTS'] = df_3a4.COMMENTS.map(lambda x: str(x)[str(x).index('|20'):] if '|20' in x else x)

    df_3a4.loc[:, 'COMMENTS'] = df_3a4.COMMENTS.map(lambda x: str(x).replace(']\n', '') if (x != np.nan and x!='') else '')

    df_3a4.loc[:, 'COMMENTS'] = df_3a4.COMMENTS.map(lambda x: str(x).replace(']', '') if (x != np.nan and x!='') else '')

    df_3a4.loc[:, 'COMMENTS']=np.where(df_3a4.COMMENTS=='nan',np.nan,df_3a4.COMMENTS)

    return df_3a4


def extract_category_from_3a4(df_3a4):
    '''
    extract the level 2 categories within brackets { } from the comments.
    :param df_3a4:
    :return:
    '''
    regex=re.compile(r'\{\D+}')

    df_3a4.loc[:, 'category_comments'] = df_3a4.COMMENTS.map(lambda x: regex.search(str(x)).group() if regex.search(str(x))!=None else np.nan)
    df_3a4.loc[:, 'category_comments']=np.where(df_3a4.category_comments.notnull(),
                                                df_3a4.category_comments.map(lambda x: str(x).replace('{','').replace('}','')),
                                                np.nan)

    return df_3a4


def find_sheet_name(outlier_elements, days_back):
    '''
    根据指定的天数差值，从outlier tracker文件中找到按日期做sheet name的sheet
    :param outlier_elements:
    :param days_back:
    :return:
    '''
    sheet_dict = {}
    for key, value in outlier_elements.items():
        fname=value[1]
        wb=openpyxl.load_workbook(fname)
        sheets=wb.sheetnames

        sheets_dated=[pd.to_datetime(x) for x in sheets]

        sheets_days_delta_vs_today=[(pd.Timestamp.now()-x).days for x in sheets_dated]

        sheets_days_delta_vs_today_vs_days_back=[x-days_back if (x-days_back)>0 else x*100 for x in sheets_days_delta_vs_today]

        min_pos=sheets_days_delta_vs_today_vs_days_back.index(min(sheets_days_delta_vs_today_vs_days_back))

        sheet_name=sheets[min_pos]

        sheet_dict[key]=sheet_name

    return sheet_dict


def calculate_outlier_changes(outlier_elements, org, df_3a4,org_name, days_back):
    '''
    Calculate the changes (number of order) from the latest df_3a4 vs. a previous outlier record from the tracker file
    :param outlier_elements:
    :param org: define which org to compare
    :param df_3a4:
    :param sht_name: old outlier sheet name in the tracker file
    :return:
    '''

    sheet_dict=find_sheet_name(outlier_elements, days_back)

    outlier_change = []
    for key, value in outlier_elements.items():
        df_old = pd.read_excel(value[1], sheet_name=sheet_dict[key])
        df_old=df_old[df_old.ORGANIZATION_CODE.isin(org_name[org])]
        if 'PO_NUMBER' in df_old.columns:
            po_old = df_old[(df_old[key] > value[0])&(df_old.ORGANIZATION_CODE.isin(org_name[org]))].PO_NUMBER.unique()
            po_new = df_3a4[(df_3a4[key] > value[0])&(df_3a4.ORGANIZATION_CODE.isin(org_name[org]))].PO_NUMBER.unique()
        else:
            po_old = df_old[(df_old[key] > value[0])&(df_old.ORGANIZATION_CODE.isin(org_name[org]))].SO_SS.unique()
            po_new = df_3a4[(df_3a4[key] > value[0])&(df_3a4.ORGANIZATION_CODE.isin(org_name[org]))].SO_SS.unique()

        po_removed = len(np.setdiff1d(po_old, po_new))
        po_added = len(np.setdiff1d(po_new, po_old))

        outlier_change.append([key + '>' + str(value[0])+'days',po_removed, po_added,len(po_new)])

    df = pd.DataFrame(outlier_change, columns=['Outlier categories','Order closed', 'Order newly added','Total open order'])

    del df_old, outlier_change
    gc.collect()

    return df


def add_outlier_to_tracker(df_3a4,outlier_elements, df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel,
                           df_partial_staged,df_missed_recommit):
    '''
    (this only happens when running APJC outlier summary)Save the outlier df to tracker file for future use. new sheet
    name is based on latest LINE_CREATION_DATE in the 3a4； only keep 14 record.
    :param df_3a4: to calculate the latest LINE_CREATION_DATE to use as sheet name
    :param outlier_elements:
    :param df_not_booked:
    :param df_not_scheduled:
    :param df_not_packed:
    :param df_aging_cancel:
    :param df_partial_staged:
    :return: None.. save the new tracker file.
    '''
    #date_mark=pd.Timestamp.now().strftime('%Y-%m-%d')
    #use the latest line creation date as the sheet_name
    date_mark=df_3a4.LINE_CREATION_DATE.sort_values().iloc[-1].strftime('%Y-%m-%d')

    for key, value,df_outlier in zip(outlier_elements.keys(),outlier_elements.values(),[df_not_booked, df_not_scheduled, df_not_packed, df_aging_cancel, df_partial_staged,df_missed_recommit]):
        df_dict = {}

        if df_outlier.shape[0]>0:
            fname=value[1]
            wb=openpyxl.load_workbook(fname)
            sheets=wb.sheetnames

            if not date_mark in sheets:
                for sheet in sheets[-14:]: # 只保留15天的记录
                    df=pd.read_excel(fname,sheet_name=sheet)

                    if df.shape[0]>0:
                        df.set_index('ORGANIZATION_CODE',inplace=True)
                    df_dict[sheet]=df.iloc[:,:4]

                df_outlier.set_index('ORGANIZATION_CODE',inplace=True)
                df_dict[date_mark]=df_outlier.iloc[:,:4]

                write_excel_file(fname, df_dict)
                # reset index, otherwise when later code call set_index(org) will have error.
                df_outlier.reset_index(inplace=True)


def get_file_from_server(path,fname):

    att_files = [(path, fname)]
    to_address=['kwang2@cisco.com']
    subject='Downloaded file from server'
    html_template='download_tracker.html'
    msg,size_over_limit =send_attachment_and_embded_image(to_address, subject, html_template,
                                               att_filenames=att_files, embeded_filenames=None,bcc=None)
    print(msg)


def del_memory(df_3a4, addr_df_summary, addr_df_dict, addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf,
                   ctb_summary_for_material):
    '''
    delete memory after run
    :return:
    '''

    del df_3a4, addr_df_summary, addr_df_dict, addr_ctb_by_org_bu, addr_ctb_by_org_bu_pf,ctb_summary_for_material

    gc.collect()

def send_downloaded_table(table_name,df,show_last,criteria_string,records_limit):
    '''
    send the downloaded file via email
    :param df:
    '''
    # Create the file
    fname=table_name + ' download.xlsx'
    df.to_excel(fname, index=False)
    full_path=os.getcwd()
    att_file = [(full_path, fname)]  # List of tuples (path, file_name)

    # Send outlier charts and spreadsheet to defined users
    to_address = ['kwang2@cisco.com']
    subject = 'Table data download: {}'.format(table_name)
    html = 'table_download.html'

    send_attachment_and_embded_image(to_address, subject, html,
                                    att_filenames=att_file,
                                    bcc = None,
                                     show_last=show_last,
                                     criteria_string=criteria_string,
                                     records_limit=records_limit)


def create_tan_asp(df_3a4):
    """
    create a pivot_table for all TAN (all options) and show corresponding unstaged qty and rev
    :param df_3a4:
    :return: pivoted df
    """
    df_asp=df_3a4.pivot_table(index=['ORGANIZATION_CODE','TAN'],values=['C_UNSTAGED_DOLLARS','C_UNSTAGED_QTY'],aggfunc=sum)
    #TODO: if same TAN appear multiple time in one PO above code may result in duplidate po_rev being added - need enhance

    df_asp.loc[:,'ASP']=df_asp.C_UNSTAGED_DOLLARS/df_asp.C_UNSTAGED_QTY
    df_asp.columns=['Total_tan_unstg_qty','Total_po_rev_unstg','TAN_ASP']
    df_asp=df_asp.applymap(float)

    df_asp.sort_values(by=['ORGANIZATION_CODE','TAN_ASP'],ascending=False,inplace=True)

    return df_asp

def redefine_addressable_flag_new(df_3a4,mfg_holds):
    '''
    Updated on Aug 26, 2020 to leveraging existing addressable definition of Y, and redefine the NO to MFG_HOLD,UNSCHEDULED,PACKED,PO_CANCELLED,NON_REVENUE
    MFG_HOLD included current MFG_HOLD col, and also the production prevention hold categories to fix some missing
    MFG_HOLD due to bug related to option PID (not showing correct MFG_HOLD when downloading option PID from 3a4)
    :param df_3a4:
    :param mfg_holds: production prevention holds set up in blg_setting
    :return:
    '''

    # Convert YES to ADDRESSABLE
    df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(df_3a4.ADDRESSABLE_FLAG=='YES',
                                                 'ADDRESSABLE',
                                                 df_3a4.ADDRESSABLE_FLAG)


    # Non_revenue orders
    df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(df_3a4.REVENUE_NON_REVENUE=='NO',
                                                  'NON_REVENUE',
                                                df_3a4.ADDRESSABLE_FLAG)

    # 如果没有LT_TARGET_FCD/TARGET_SSD/CURRENT_FCD_NBD_DATE,则作如下处理 - 可能是没有schedule或缺失Target LT date or Target SSD
    df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(df_3a4.CURRENT_FCD_NBD_DATE.isnull(),
                                                 'UNSCHEDULED',
                                                 df_3a4.ADDRESSABLE_FLAG)
    # update MFG_HOLD
    if 'OPTION_NUMBER' not in df_3a4.columns:
        df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(df_3a4.MFG_HOLD=='Y',
                                                     'MFG_HOLD',
                                                     df_3a4.ADDRESSABLE_FLAG)
    else:
        # Get hold PO from 'MFG_HOLD' label (this also capture some hold on option PID)
        mfg_hold_po = df_3a4[
            df_3a4.MFG_HOLD == 'Y'].PO_NUMBER.unique()  # do this to capture all MFG hold order and add it on all PIDs below

        # Get hold PO based on mfg_holds
        df_hold = df_3a4[df_3a4.ORDER_HOLDS.notnull()].copy()
        if df_hold.shape[0] > 0:  # in case there is not any hold orders
            df_hold.loc[:, 'mfg_hold'] = np.nan
            for hold in mfg_holds:
                df_hold.loc[:, 'mfg_hold'] =np.where(df_hold.ORDER_HOLDS.str.contains(hold, case=False),
                                                    'YES',
                                                     df_hold.mfg_hold)

            mfg_hold_po_other = df_hold[df_hold.mfg_hold == 'YES'].PO_NUMBER.unique()

            # combined po list and redefine the addressable flag
            mfg_hold_po = mfg_hold_po.tolist() + mfg_hold_po_other.tolist()

        df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(df_3a4.PO_NUMBER.isin(mfg_hold_po),
                                                     'MFG_HOLD',
                                                     df_3a4.ADDRESSABLE_FLAG)

    # redefine cancellation order to PO_CANCELLED - put this as the last
    # take extra step below to have "PO_CANCELLED" shown on all options as well
    df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where(
        (df_3a4.ORDER_HOLDS.str.contains('Cancellation', case=False)) & (df_3a4.ORDER_HOLDS.notnull()),
        'PO_CANCELLED',
        df_3a4.ADDRESSABLE_FLAG)
    # further do below if there is option in 3a4
    if 'OPTION_NUMBER' in df_3a4.columns:
        cancel_po=df_3a4[df_3a4.ADDRESSABLE_FLAG=='PO_CANCELLED'].PO_NUMBER.unique()
        df_3a4.loc[:, 'ADDRESSABLE_FLAG']=np.where(df_3a4.PO_NUMBER.isin(cancel_po),
                                               'PO_CANCELLED',
                                               df_3a4.ADDRESSABLE_FLAG)


    """ Discarded as some of the packed are under addressable, discard this so we can be 100% consistent with UOV addressable YES
    # change packed orders flag to PACKED IF NOT cancelled
    df_3a4.loc[:, 'ADDRESSABLE_FLAG'] = np.where((df_3a4.PACKOUT_QUANTITY == 'Packout Completed') & (df_3a4.ADDRESSABLE_FLAG!='PO_CANCELLED'),
                                                           'PACKED',
                                                           df_3a4.ADDRESSABLE_FLAG)
    """

    # OTHER non-addressable
    df_3a4.loc[:, 'ADDRESSABLE_FLAG']=np.where(df_3a4.ADDRESSABLE_FLAG=='NO',
                                               'NOT_ADDRESSABLE',
                                               df_3a4.ADDRESSABLE_FLAG)

    return df_3a4


def test_sending_dashboard():
    to_address='kwang2@cisco.com'
    subject='dashboard'
    html_template='dashboard-example.html'
    msg, size_over_limit = send_attachment_and_embded_image(to_address, subject, html_template,
                                                            att_filenames=None, embeded_filenames=None,bcc = None)


if __name__ == '__main__':
    """
    to_num = '+8618665932236'
    message = 'this is a test message from twilio'
    sms = SendSms()
    sms.send_sms(message, to_num)
    """
    login_user='kw'
    backup_day='Sunday'
    download_and_send_tracker_as_backup(backup_day,login_user)